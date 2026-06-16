import openpyxl
import sys
import io
import os

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

def get_rows(fpath, sheet_name=None, max_rows=100, max_cols=None):
    if not os.path.exists(fpath):
        return [[f"ERROR: File not found {fpath}"]]
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active
        rows_data = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if r_idx >= max_rows: break
            if any(c is not None for c in row):
                cleaned_row = [str(c).strip().replace('\n', ' ') if c is not None else '' for c in row]
                if max_cols: cleaned_row = cleaned_row[:max_cols]
                rows_data.append(cleaned_row)
        return rows_data
    except Exception as e:
        return [[f"ERROR reading {fpath}: {e}"]]

print('--- CLIENT REQUIREMENTS (Reports 1) ---')
f1 = r'd:\Steamee\Reports 1 , detailed store wise.xlsx'
for sheet in ['Sheet1', 'Sheet2']:
    print(f'Sheet: {sheet}')
    for r in get_rows(f1, sheet):
        print(' | '.join(r))

print('\n--- OUR KPI FRAMEWORK (v3) ---')
f2 = r'd:\Steamee\STEAMEE_KPI_Change_Log_v3.xlsx'
for r in get_rows(f2, 'KPI MASTER LIST', max_cols=3):
    print(' | '.join(r))

print('\n--- BALA API LIST ---')
f3 = r'd:\Steamee\steame_bala_sr_final.xlsx'
for r in get_rows(f3, max_cols=5):
    print(' | '.join(r))
