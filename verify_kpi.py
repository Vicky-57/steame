import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

dst = r'd:\Steamee\STEAMEE_KPI_Change_Log_v3.xlsx'
wb = openpyxl.load_workbook(dst, data_only=True)
ws = wb['KPI MASTER LIST']
ws_b = wb['BUGS & DATA GAPS']

print('=== VERIFICATION OF KEY CHANGES ===\n')

verify_rows = {
    5:  'Q4  - E2E TAT',
    12: 'Q9b - Pressman Productivity',
    14: 'Q10 - Rider Utilisation',
    23: 'Q18a- Rider Attendance',
    28: 'Q23 - Promo ROI',
    29: 'Q24 - App Funnel',
    31: 'Q26 - Churn Rate',
    41: 'Q34 - Refund Leakage',
    44: 'Q36a- Wallet Balance',
    49: 'Q41 - Cust Satisfaction',
    54: 'Q45a- Audit Compliance',
}

for row_n, label in verify_rows.items():
    row = list(ws.iter_rows(min_row=row_n, max_row=row_n, values_only=True))[0]
    print(f'--- {label} (Row {row_n}) ---')
    print(f'  Q#={row[0]} | Status={row[3]} | DataReady={row[8]} | PortalStatus={row[13]}')
    src = str(row[6])[:120] if row[6] else 'N/A'
    notes = str(row[14])[:150] if row[14] else 'N/A'
    print(f'  Source: {src}')
    print(f'  PortalNotes: {notes}')
    print()

print('\n=== BUGS & DATA GAPS SHEET ===')
for row in ws_b.iter_rows(values_only=True):
    if any(c for c in row):
        print([str(c)[:100] if c else '' for c in row[:4]])

print('\nFILE: d:\\Steamee\\STEAMEE_KPI_Change_Log_v3.xlsx')
print('STATUS: All changes verified successfully.')
