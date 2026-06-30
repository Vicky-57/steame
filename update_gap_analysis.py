# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

f_reports1 = r"d:\Steamee\Reports 1 , detailed store wise.xlsx"
f_gap = r"d:\Steamee\STEAMEE_Gap_Analysis.xlsx"

# Standard Colors
NAVY_FILL = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
WHITE_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
REGULAR_FONT = Font(name="Calibri", size=11)

# Status Fills & Fonts
GREEN_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
GREEN_FONT = Font(name="Calibri", size=11, color="155724", bold=True)

YELLOW_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
YELLOW_FONT = Font(name="Calibri", size=11, color="856404", bold=True)

RED_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
RED_FONT = Font(name="Calibri", size=11, color="721C24", bold=True)

# Borders
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

def apply_auto_width_and_styles(ws):
    ws.views.sheetView[0].showGridLines = True
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value and not str(cell.value).startswith('='):
                lines = str(cell.value).split('\n')
                for line in lines:
                    max_len = max(max_len, len(line))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 55)

# Load existing base gap analysis rows
wb_reports1 = openpyxl.load_workbook(f_reports1, data_only=True)
ws_s2 = wb_reports1['Sheet2']

wb_gap = openpyxl.Workbook()
ws_gap = wb_gap.active
ws_gap.title = "GAP ANALYSIS"

headers_gap = [
    "Client Expectation (Reports 1)",
    "Category",
    "Status",
    "KPI Mapping in v3",
    "Technical Source / Solution",
    "What is Covered / Plain-English Note",
    "What is Missing / Data Gaps",
    "New API Needed?",
    "Implementation Priority"
]
ws_gap.append(headers_gap)

# Style Header
for col_idx, header in enumerate(headers_gap, 1):
    cell = ws_gap.cell(row=1, column=col_idx)
    cell.fill = NAVY_FILL
    cell.font = WHITE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws_gap.row_dimensions[1].height = 30

current_category = ""
row_data_list = []

for r_idx, row in enumerate(ws_s2.iter_rows(values_only=True)):
    if r_idx == 0: continue
    cat = row[0]
    metric = row[1]
    definition = row[2] if len(row) > 2 and row[2] is not None else ""
    
    if cat: current_category = str(cat).strip()
    if not metric: continue
    
    m_str = str(metric).strip()
    def_str = str(definition).strip()
    
    status = "Covered"
    kpi_map = "Q9"
    source = "orders table"
    covered_note = "Tracked inside the core operational order states dashboard."
    missing_note = "None. Handled by existing data schema."
    new_api = "No"
    priority = "P0 - CRITICAL"
    
    m_lower = m_str.lower()
    c_lower = current_category.lower()
    
    # Standard rules
    if "stock" in c_lower or "stock" in m_lower:
        status = "Gap"
        kpi_map = "N/A"
        source = "None - Needs API"
        covered_note = "No stock metrics are currently in v3."
        missing_note = "Missing daily opening stock, closing stock, and receipts database logs."
        new_api = "Yes - /api/stock-inventory"
        priority = "P0 - CRITICAL"
    elif "machine" in c_lower or "cleanliness" in m_lower or "downtime" in m_lower:
        status = "Gap"
        kpi_map = "N/A"
        source = "None - Needs API"
        covered_note = "No machine metrics are currently in v3."
        missing_note = "Missing machine service cost tracking, store maintenance costs, and cleanliness score."
        new_api = "Yes - /api/machine-maintenance"
        priority = "P1 - HIGH"
    elif "pincode" in c_lower or "pincode" in m_lower:
        status = "Gap"
        kpi_map = "N/A"
        source = "None - Needs API"
        covered_note = "Currently tracked only at zone/store level."
        missing_note = "Need pincode field mapped in orders and users APIs to track sales, cohorts, and slot trends."
        new_api = "Yes - /api/pincode-analytics"
        priority = "P1 - HIGH"
    elif "marketing" in c_lower or "platform" in m_lower or " CAC " in m_str or "cac" in m_lower or "attribut" in m_lower or "aquired through" in m_lower:
        status = "Gap"
        kpi_map = "Q22"
        source = "None - Needs API"
        covered_note = "Q22 (CAC) is flagged in v3 but data is NOT ready due to missing database tables (G1)."
        missing_note = "No digital marketing spend tracking exists. Blocked by missing marketing spend table."
        new_api = "Yes - /api/marketing-attribution"
        priority = "P2 - PHASE 2"
    elif "delayed order" in m_lower or "24hrs" in m_lower:
        status = "Gap"
        kpi_map = "N/A"
        source = "None - Needs API"
        covered_note = "SLA tracking exists (Q1, Q2, Q3) but no specific indicator flags orders >24h in ironing/delivery."
        missing_note = "Requires store manager real-time flag for bags delayed beyond 24 hours."
        new_api = "Yes - /api/delayed-orders"
        priority = "P0 - CRITICAL"
    elif "bag" in m_lower and "inventory" in c_lower or "received bags" in m_lower or "pending bag" in m_lower:
        status = "Gap"
        kpi_map = "N/A"
        source = "None - Needs API"
        covered_note = "Smart bag utilization (Q25) is tracked but not the physical bag inventory."
        missing_note = "Needs bag inventory scans inside stores (received list vs delivery-pending bags)."
        new_api = "Yes - /api/order-inventory"
        priority = "P0 - CRITICAL"
    elif "invoice" in m_lower or "accounting" in c_lower or "bills" in m_lower:
        status = "Gap"
        kpi_map = "N/A"
        source = "None - Needs API"
        covered_note = "Sales and COD collections tracked, but not invoicing documents."
        missing_note = "Need backend automated billing and invoice PDF generation."
        new_api = "Yes - /api/invoicing"
        priority = "P2 - PHASE 2"
    elif "retention" in m_lower or "cohort" in m_lower:
        status = "Covered"
        kpi_map = "Q26"
        source = "cohort_analytics API"
        covered_note = "Cohort Retention rates over months/year tracked dynamically."
        missing_note = "None."
        new_api = "No"
        priority = "P0 - CRITICAL"
    elif "complaint" in c_lower or "complaint" in m_lower:
        if "48hrs" in m_lower or "beyond" in m_lower:
            status = "Partial"
            kpi_map = "Q42"
            source = "complaint_refund API"
            covered_note = "Tracks support resolution speed. Duration can be calculated from complaint creation date."
            missing_note = "Requires a Metabase timestamp duration logic, not a default card yet."
            new_api = "No"
            priority = "P1 - HIGH"
        else:
            status = "Covered"
            kpi_map = "Q42"
            source = "complaint_refund API"
            covered_note = "Fully covers placed, open, closed, in-process states and types."
            missing_note = "None."
            new_api = "No"
            priority = "P0 - CRITICAL"
    elif "sales" in c_lower or "sales" in m_lower or "revenue" in m_lower or "cash" in m_lower or "online" in m_lower or "charges" in m_lower:
        status = "Covered"
        if "avg order value" in m_lower or "avg order size" in m_lower:
            kpi_map = "Q30"
            source = "v_daily_store_summary"
            covered_note = "Calculated as Gross Revenue / Completed Orders."
        elif "discount" in m_lower or "offer" in m_lower:
            kpi_map = "Q23"
            source = "orders table"
            covered_note = "Voucher and Promo discounts tracked directly from order fields."
        elif "cash" in m_lower or "online" in m_lower or "cod" in m_lower:
            kpi_map = "Q31"
            source = "orders table"
            covered_note = "Fully mapped using payment mode ('Cash' instead of COD, online options)."
        elif "return orders" in m_lower:
            status = "Partial"
            kpi_map = "Q17"
            source = "orders table"
            covered_note = "Logistics cost tracked. Isolated return fees are partially covered."
            missing_note = "Needs dedicated field for return logistics fee in DB."
            priority = "P1 - HIGH"
        else:
            kpi_map = "Q28"
            source = "v_daily_store_summary"
            covered_note = "Gross Revenue tracked per store, zone, and day."
        new_api = "No"
    elif "customer" in c_lower or "user" in m_lower or "active" in m_lower or "churn" in m_lower:
        if "download" in m_lower or "new" in m_lower:
            kpi_map = "Q20"
            source = "user_analytics API"
            covered_note = "Tracks new registrations and app funnel conversion."
        elif "inactive" in m_lower or "churn" in m_lower:
            kpi_map = "Q26"
            source = "cohort_analytics API"
            covered_note = "Inactive customers and churn counts mapped under cohort analysis."
        else:
            kpi_map = "Q19"
            source = "user_analytics API"
            covered_note = "Weekly/Monthly Active Users (WAU/MAU) fully tracked."
        new_api = "No"
        priority = "P0 - CRITICAL"
    elif "ironer" in c_lower or "pressman" in m_lower or "payout" in m_lower or "clothes" in m_lower:
        if "payout" in m_lower:
            status = "Partial"
            kpi_map = "Q9b"
            source = "ironer_performance_analytics + payouts"
            covered_note = "Pressman individual payout calculated by clothes processed."
            missing_note = "Payout report structure needs formatting on database side (G4)."
            new_api = "No"
            priority = "P1 - HIGH"
        elif "kyc" in m_lower:
            status = "Gap"
            kpi_map = "N/A"
            source = "None - Needs KYC verification system"
            covered_note = "KPI framework covers pressman performance but not onboarding documents."
            missing_note = "Missing pressman KYC status columns."
            new_api = "No"
            priority = "P2 - PHASE 2"
        elif "process follow" in m_lower or "process lag" in m_lower:
            status = "Gap"
            kpi_map = "N/A"
            source = "None - Needs stage scan events"
            covered_note = "KPI framework tracks defects/complaints but not the individual process steps."
            missing_note = "Needs barcode scans at every micro-stage in the workshop."
            new_api = "No"
            priority = "P2 - PHASE 2"
        else:
            kpi_map = "Q9b"
            source = "ironer_performance_analytics API"
            covered_note = "Tracks pressman performance: ironed clothes count, defect rate, and efficiency."
            new_api = "No"
            priority = "P0 - CRITICAL"
    elif "logistics" in c_lower or "rider" in m_lower or "route" in m_lower or "heat mapping" in m_lower:
        if "route" in m_lower or "batching" in m_lower or "heat mapping" in m_lower:
            status = "Gap"
            kpi_map = "N/A"
            source = "None - Needs GPS tracking"
            covered_note = "Logistics tracked by delivery zones and orders, not by active map coordinates."
            missing_note = "Requires coordinate tracking from rider app and mapping SDK integrations."
            new_api = "No"
            priority = "P2 - PHASE 2"
        elif "reconsilation" in m_lower or "cash collection" in m_lower:
            status = "Partial"
            kpi_map = "Q13"
            source = "rider_performance_analytics + COD"
            covered_note = "Rider collection rate tracked, but not automated deposit reconciliation."
            missing_note = "Reconciliation workflow of rider cash matches needs database structure."
            new_api = "No"
            priority = "P1 - HIGH"
        else:
            kpi_map = "Q10"
            source = "rider_performance_analytics API"
            covered_note = "Rider utilization, deliveries per day, SLAs, and order details."
            new_api = "No"
            priority = "P0 - CRITICAL"
    elif "order status" in c_lower or "pickup" in m_lower or "delivery" in m_lower or "sla" in m_lower:
        if "pickup" in m_lower:
            kpi_map = "Q1"
        else:
            kpi_map = "Q3"
        source = "v_sla_tracking"
        covered_note = "SLA on-time pick up and delivery percentages are fully tracked."
        new_api = "No"
        priority = "P0 - CRITICAL"
    elif "slot" in c_lower or "slot" in m_lower:
        kpi_map = "Q9a"
        source = "orders table"
        covered_note = "Tracks slot capacities vs scheduled demand to optimize operations."
        new_api = "No"
        priority = "P0 - CRITICAL"
    elif "roles" in m_lower or "permission" in m_lower:
        kpi_map = "N/A"
        source = "Metabase permissions layer"
        covered_note = "Platform-level user roles (admin, store manager, logistics lead) govern access."
        missing_note = "None."
        new_api = "No"
        priority = "P0 - CRITICAL"
    elif "trigger" in m_lower or "automated" in m_lower or "auto message" in m_lower or "retargetting" in c_lower:
        status = "Gap"
        kpi_map = "N/A"
        source = "None - Needs CRM integration"
        covered_note = "Metabase displays active/inactive splits but cannot trigger actions."
        missing_note = "Requires integration with MoEngage or Clevertap for event-driven message triggers."
        new_api = "No"
        priority = "P2 - PHASE 2"
        
    row_data_list.append([m_str, current_category, status, kpi_map, source, covered_note, missing_note, new_api, priority])

# Add the new client-feedback mapping entries
new_kpi_additions = [
    ["Ironer Performance Summary Table", "Store Operations", "Covered", "Q9b", "ironer-analytics", "Tracks ironer levels (lv1, lv2, trainee), active leaderboards, present counts, and attendance stats.", "None.", "No", "P0 - CRITICAL"],
    ["Staff & Rider Attendance Shift Tracker", "Store Operations", "Covered", "Q18a", "attendance-analytics", "Logs check-in/out times, working hours, shift name, and durations.", "None.", "No", "P0 - CRITICAL"],
    ["Today's Completed Sales vs Pending Sales", "Sales & Finance", "Covered", "Q28", "orders table / v_daily_store_summary", "Splits billed sales from orders currently in processing queue.", "None.", "No", "P0 - CRITICAL"],
    ["Sales Volume vs. Order Size Analysis", "Sales & Finance", "Covered", "Q28, Q30", "orders / /api/order_revenue_trend", "Tracks relationship between sales, order counts, and average order value.", "None.", "No", "P0 - CRITICAL"],
    ["Pending Orders Sales Contribution", "Sales & Finance", "Covered", "Q9, Q28", "orders / /api/orders/list", "Tracks estimated revenue value of active orders currently in processing queue.", "None.", "No", "P0 - CRITICAL"],
    ["Customer Churn Drop-Off Inactivity Bands", "User Analytics", "Covered", "Q26", "cohort-analytics", "Groups inactive/churned customer volumes specifically into 30-45d, 46-60d, 61-90d, 91-200d, and 200+d bands.", "None.", "No", "P0 - CRITICAL"],
    ["User Journey & Stage Migration Timelines", "User Analytics", "Covered", "N/A", "user-analytics", "Measures average days customer takes to move from Onboarded -> Active -> Value -> Super Value.", "None.", "No", "P1 - HIGH"],
    ["Stage Conversion Funnel storewise", "User Analytics", "Covered", "Q19, Q26", "user-analytics / /api/getUserSegments", "Splits customer counts by lifecycle stage (Onboarded, Churned, Active, Inactive) per store.", "None.", "No", "P0 - CRITICAL"],
    ["Persona Complaint & Ticket Breakdown", "User Analytics", "Covered", "Q42", "complaintListApi / /api/get_complaint_persona_trend", "Correlates complaint and ticket volumes split by customer loyalty tier.", "None.", "No", "P0 - CRITICAL"],
    ["Stage-wise Pending TAT Queues", "Store Operations", "Covered", "Q4", "v_sla_tracking", "Breaks down pending hours spent specifically in Pickup stage, Ironing stage, and Delivery stage.", "None.", "No", "P0 - CRITICAL"],
    ["Avg Store-wise Missed TAT SLA", "Store Operations", "Covered", "Q4", "v_sla_tracking / /api/get_store_tat_analysis", "Tracks total volume and percentage of orders that missed their SLA window per store.", "None.", "No", "P0 - CRITICAL"],
    ["Store-wise Ironer Performance Matrix", "Store Operations", "Covered", "Q9b, Q18a", "ironer-performance-analytics", "Details individual ironer quantity, quality, speed, complaints, attendance, and calculated earnings.", "None.", "No", "P0 - CRITICAL"],
    ["Store-wise Rider Performance Matrix", "Store Operations", "Covered", "Q10, Q12, Q13", "rider-performance-analytics", "Details individual rider quantity, speed, complaints, attendance, and payouts.", "None.", "No", "P0 - CRITICAL"],
    ["Asset Repair & Maintenance Timeline Log", "Store Operations", "Gap", "N/A", "None - Needs Asset logs", "Logs repair timelines, breakdown downtime hours, and service costs for store machines.", "Missing database table for asset repair tracking.", "Yes - /api/machine-maintenance", "P1 - HIGH"],
    ["Inventory Usage Forecasting Ledger", "Store Operations", "Gap", "N/A", "None - Needs stock logs", "Tracks daily usage of chemicals/packaging and forecasts replenishment dates based on usage speed.", "Missing stock/chemical ledger logs.", "Yes - /api/stock-inventory", "P1 - HIGH"],
    ["Operations Manager Performance Scorecard", "Store Operations", "Covered", "Q40", "v_daily_store_summary", "Ranks store managers on store SLA, sales growth, user churn, and team complaints.", "None.", "No", "P0 - CRITICAL"],
    ["Smart Bag Cycle & Lifespan Tracker", "Store Operations", "Covered", "Q25", "orders table / /api/user-analytics", "Logs average bags allocated per customer and rotation frequency count.", "None.", "No", "P0 - CRITICAL"],
    ["First-Service SLA Defection Funnel", "Store Operations", "Covered", "Q24, Q26", "user-analytics / /api/cohort-analytics", "Tracks churn rates specifically following delayed first orders or clothing damage.", "None.", "No", "P0 - CRITICAL"],
    ["Rider Shift Logs & Duration Details", "Rider Logistics", "Covered", "Q18a", "attendance-analytics", "Tracks rider login, logout, session active duration minutes, and late clock-in counts.", "None.", "No", "P0 - CRITICAL"],
    ["door-to-door Logistics Fulfilment Timeline Tracker", "Rider Logistics", "Covered", "Q18", "rider-performance-analytics", "Tracks elapsed minutes from order assign -> arrival -> wait -> transit -> drop-off.", "None.", "No", "P0 - CRITICAL"],
    ["Logistics Performance Breakdown Matrix", "Rider Logistics", "Covered", "Q10, Q11, Q14", "rider-performance-analytics", "Breaks down logistics acceptance, cancellations, and SLA breaches by zone, rider, and slot.", "None.", "No", "P0 - CRITICAL"],
    ["Prepaid Wallet Recharge & Consumption Ledger", "Finance & Revenue", "Covered", "Q36a", "wallet-analytics", "Tracks cash wallet, reward wallet, referral wallet additions, recharges, and order consumption.", "None.", "No", "P0 - CRITICAL"],
    ["User-Type Customer Lifetime Value (LTV)", "Growth & Marketing", "Covered", "Q38", "user-analytics / /api/user-analytics", "Tracks customer LTV split specifically by user segments (Onboarded, Active, Value, VIP).", "None.", "No", "P0 - CRITICAL"],
    ["Complaint vs Churn vs RTO Correlation", "Quality & Complaint", "Covered", "Q42, Q26", "complaintListApi / cohort-analytics", "Correlates complaints and RTO returned orders with monthly customer churn rates.", "None.", "No", "P0 - CRITICAL"],
    ["Customer Service Executive Performance Scorecard", "Quality & Complaint", "Covered", "N/A", "complaintListApi / /api/get_staff_detailed_performance", "Ranks support agents on resolution TAT, count and cost of refunds given, and customer conversion.", "None.", "No", "P1 - HIGH"],
    # LATEST ADDITIONS (DELIVERY PARTNER PERFORMANCE, EXPIRED WALLETS, EXPENSES, MASTERS CONFIG & LISTS)
    ["Delivery Partner Performance Grid (3P Fleets)", "Rider Logistics", "Covered", "Q10, Q12", "cod-analytics / /api/cod-analytics", "Monitors Pidge, Porter, Dunzo delivery partner SLA compliances, cancels, and cost-per-completed-task.", "None.", "No", "P0 - CRITICAL"],
    ["Money Wallet Utilised vs Unutilised Split", "Finance & Revenue", "Covered", "Q36a", "wallet-analytics / /api/wallet-analytics", "Tracks utilized vs unutilized customer money wallet balances.", "None.", "No", "P0 - CRITICAL"],
    ["Expired Reward Wallet Value", "Finance & Revenue", "Covered", "Q36a", "wallet-analytics / /api/wallet-analytics", "Tracks reward/promotional coins that expired without being used.", "None.", "No", "P0 - CRITICAL"],
    ["Store Expense Management Ledger", "Finance & Revenue", "Gap", "N/A", "None - Needs expense logs", "Tracks daily chemical, pressman, rider, and utility store expense transactions.", "Missing database table for store-level expense logging.", "Yes - /api/stock-inventory", "P1 - HIGH"],
    ["Masters Configuration Summary", "Store Operations", "Covered", "N/A", "None - Masters screens", "Lists all 15 configuration master tables (Community Pricing, Delivery, Price, Pincode, etc.) exactly as is.", "None.", "No", "P0 - CRITICAL"],
    ["Wardrobe (Garment Proof Logs)", "Store Operations", "Covered", "N/A", "None - Operational lists", "Details order garment proof photos uploaded by vendors for quality logs.", "None.", "No", "P0 - CRITICAL"],
    ["Rack Management Layout Grid", "Store Operations", "Covered", "N/A", "None - Operational lists", "Maps visual occupied vs empty racks with order bag barcode tags.", "None.", "No", "P0 - CRITICAL"],
    ["Customer Support Chat Room", "Quality & Complaint", "Covered", "N/A", "None - Operational lists", "Direct chat logs with direct users, store managers, and riders.", "None.", "No", "P0 - CRITICAL"],
    ["Zonewise Report List Grid", "Store Operations", "Covered", "N/A", "None - Operational lists", "S.No, Customer Name, Mobile, OrderId, zone name, Cloth Count, Date.", "None.", "No", "P0 - CRITICAL"],
    ["Offer Report List Grid", "Store Operations", "Covered", "N/A", "None - Operational lists", "Used At, Promocode, User Name, Order ID, Order Amount, Promo Discount, Payable, Zone.", "None.", "No", "P0 - CRITICAL"],
    ["Vendor Summary Report Grid", "Store Operations", "Covered", "N/A", "None - Operational lists", "Total orders, Unit, Vendor, Total Ironers, Cloth counts, Queues, Completed.", "None.", "No", "P0 - CRITICAL"],
    # NEW DETAILED TABLES INCORPORATION
    ["Asset Category & Product Types Configuration", "Store Operations", "Gap", "N/A", "None - Needs Asset database", "Structure materials and table assets categorization with category codes.", "Missing database tables for physical asset definitions.", "Yes - /api/asset-management", "P1 - HIGH"],
    ["Asset Entry & Status Management", "Store Operations", "Gap", "N/A", "None - Needs Asset database", "Track active assets, purchase date, serial numbers, warranty, AMC and Out-for-service transitions.", "Missing database tables for physical asset logging.", "Yes - /api/asset-management", "P1 - HIGH"],
    ["Asset Service Report Log", "Store Operations", "Gap", "N/A", "None - Needs Asset database", "Logs down-time, service start/end dates, replacement asset codes, purposes and remarks.", "Missing database tables for asset servicing.", "Yes - /api/asset-management", "P1 - HIGH"],
    ["Inventory Warehouse Setup Master", "Store Operations", "Gap", "N/A", "None - Needs Inventory database", "Setup supply warehouses, locations, and manager mappings.", "Missing inventory warehouse database tables.", "Yes - /api/inventory-management", "P1 - HIGH"],
    ["Inventory Stock & Safety Thresholds", "Store Operations", "Gap", "N/A", "None - Needs Inventory database", "Warehouse, store, and live stocks levels tracking with safety thresholds for buttersheet, bag, and collar cards.", "Missing safety stock tracking systems.", "Yes - /api/inventory-management", "P1 - HIGH"],
    ["Inventory Store & Warehouse Transfers Requests", "Store Operations", "Gap", "N/A", "None - Needs Inventory database", "Request and dispatch flow of items from warehouses to stores.", "Missing inventory transfer log tables.", "Yes - /api/inventory-management", "P1 - HIGH"],
    ["Expense Entry & Config Logs", "Finance & Revenue", "Gap", "N/A", "None - Needs Expense database", "Expense category, sub-category, payment modes setup, and daily expense entry log.", "Missing expense and payout ledger tables.", "Yes - /api/expense-management", "P1 - HIGH"],
    ["Expense Analytics & Report Ledger", "Finance & Revenue", "Gap", "N/A", "None - Needs Expense database", "CFO dashboard comparing income, net margin, promotional, loyalty, and voucher discounts spend.", "Missing integrated expense analytics reporting systems.", "Yes - /api/expense-analytics", "P1 - HIGH"],
    ["Manual Order Form", "Store Operations", "Covered", "N/A", "orders table", "Create manual orders, cloth orders, or bag orders for walk-ins with mobile lookup.", "None.", "No", "P0 - CRITICAL"],
    ["Order List Table (Comprehensive Statuses)", "Store Operations", "Covered", "Q9", "orders table / /api/orders/list", "Reconcile order counts across 20 distinct states with active TAT timers.", "None.", "No", "P0 - CRITICAL"],
    ["Scheduled Orders & Pickup Hold Dashboard", "Store Operations", "Covered", "Q9a", "orders table", "Manage pickup hold dashboards, delivery holds, standard holds, and send selected to partner.", "None.", "No", "P0 - CRITICAL"],
    ["Onhold Orders Alerts & Duration List", "Store Operations", "Covered", "Q9", "orders table", "Tracks total on hold orders, grouping by Normal (<=1d), Warning (2-4d), and Critical (>4d).", "None.", "No", "P0 - CRITICAL"],
    ["Orders Bulk Transfer Manager", "Store Operations", "Gap", "N/A", "None - Needs transfer logs", "Interface to bulk-assign orders to different stores or processing units.", "Missing bulk transfer audit tracking in database.", "Yes - /api/order-transfers", "P1 - HIGH"],
    ["Batch Transfers List Report", "Store Operations", "Gap", "N/A", "None - Needs transfer logs", "View, filter, and print batch transfer lists with Origin Zone, Processing Zone, and Order/Bag counts.", "Missing batch transfer log tracking.", "Yes - /api/order-transfers", "P1 - HIGH"],
    ["Telecalling Status & Action Taken Setup", "Growth & Marketing", "Gap", "N/A", "None - Needs Telecaller database", "Configure telecall pipelines (Prospect, Warm, Disqualified, RNR, Ringing, etc.) and actions taken (Follow-up scheduled, App explained).", "Missing telecalling pipeline tables.", "Yes - /api/telecall-leads", "P2 - PHASE 2"],
    ["TeleCall History & Call Logs Report", "Growth & Marketing", "Gap", "N/A", "None - Needs Telecaller database", "Details callers register dates, follow-up times, priorities, disqualification reasons, and remarks.", "Missing lead history call logging.", "Yes - /api/telecall-leads", "P2 - PHASE 2"],
    ["GST Tax Master Configuration", "Store Operations", "Covered", "N/A", "orders table / settings", "Configure GST Name, GST Percent, and Active Status dropdowns.", "None.", "No", "P0 - CRITICAL"]
]

row_data_list.extend(new_kpi_additions)

# ── DYNAMIC MARKETING REQS IN GAP ANALYSIS ───────────────────────────────────
try:
    mktg_wb = openpyxl.load_workbook(r"d:\Steamee\Steamee_Marketing_Dashboard_Requirements.xlsx", data_only=True)
    mktg_ws = mktg_wb["Dashboard Specs"]

    mktg_gap_additions = []
    for mktg_ri in range(4, mktg_ws.max_row + 1):
        row_vals = [cell.value for cell in mktg_ws[mktg_ri]]
        if not any(row_vals): continue
        m_id, name, cat, priority, obj, metrics, dims, gran, sources, visuals, dod, deps = row_vals[:12]
        
        is_moengage = 'moengage' in str(sources).lower() or 'app analytics' in str(sources).lower() or 'app_open' in str(sources).lower()
        
        # 1. Option 1: Portal + MoEngage
        status_p = "Covered" if is_moengage else "Gap"
        source_p = "MoEngage Events" if is_moengage else "Portal Backend APIs"
        covered_p = f"Option 1 (Portal): {metrics if metrics else ''}"
        missing_p = "None (MoEngage tracked)" if is_moengage else f"Requires custom Portal DB endpoints and frontend UI for: {name}."
        new_api_p = "No" if is_moengage else "Yes (custom API)"
        
        mktg_gap_additions.append([
            f"[Portal] [{m_id}] {name}",
            f"Marketing - {cat}",
            status_p,
            f"{m_id}",
            source_p,
            covered_p,
            missing_p,
            new_api_p,
            priority
        ])

        # 2. Option 2: Metabase + MoEngage
        status_m = "Covered"
        source_m = "MoEngage Events" if is_moengage else "Metabase SQL Query / Postgres"
        covered_m = f"Option 2 (Metabase): {metrics if metrics else ''}"
        missing_m = "None (MoEngage tracked)" if is_moengage else "None. Directly query Postgres read-replica."
        new_api_m = "No"
        
        mktg_gap_additions.append([
            f"[Metabase] [{m_id}] {name}",
            f"Marketing - {cat}",
            status_m,
            f"{m_id}",
            source_m,
            covered_m,
            missing_m,
            new_api_m,
            priority
        ])

    row_data_list.extend(mktg_gap_additions)
except Exception as ex:
    print("⚠️ Warning generating marketing gaps:", ex)

for row_idx, r_val in enumerate(row_data_list, 2):
    ws_gap.append(r_val)
    # Styles
    for col_idx in range(1, len(headers_gap) + 1):
        cell = ws_gap.cell(row=row_idx, column=col_idx)
        cell.font = REGULAR_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        
        # Color status column (Col 3)
        if col_idx == 3:
            val = str(cell.value)
            if val == "Covered":
                cell.fill = GREEN_FILL
                cell.font = GREEN_FONT
            elif val == "Partial":
                cell.fill = YELLOW_FILL
                cell.font = YELLOW_FONT
            else:
                cell.fill = RED_FILL
                cell.font = RED_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Priority column (Col 9)
        if col_idx == 9:
            val = str(cell.value)
            if "P0" in val:
                cell.fill = RED_FILL
                cell.font = RED_FONT
            elif "P1" in val:
                cell.fill = PatternFill(start_color="FFE8D6", end_color="FFE8D6", fill_type="solid")
                cell.font = Font(name="Calibri", size=11, color="D35400", bold=True)
            else:
                cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
                cell.font = Font(name="Calibri", size=11, color="495057", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # API Needed Column (Col 8)
        if col_idx == 8:
            val = str(cell.value)
            if val.startswith("Yes"):
                cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                cell.font = RED_FONT
            else:
                cell.fill = GREEN_FILL
                cell.font = GREEN_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_gap.row_dimensions[row_idx].height = 42

apply_auto_width_and_styles(ws_gap)
wb_gap.save(f_gap)
print("STEAMEE_Gap_Analysis.xlsx updated and synchronized successfully.")
