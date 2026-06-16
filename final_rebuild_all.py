# -*- coding: utf-8 -*-
"""
STEAMEE — Final Rebuild Script
Regenerates:
  1. STEAMEE_Client_Dashboard_Spec_v2.xlsx   (11 sheets, complete detail)
  2. STEAMEE_Gap_Analysis.xlsx               (all 148+ rows with colour-coding)
  3. STEAMEE_KPI_Change_Log_v3.xlsx          (SUMMARY counts updated)
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

# ── COLOUR PALETTE ─────────────────────────────────────────────────────────────
NAVY       = "1B365D"
TEAL       = "0D7377"
GOLD       = "E8A020"
WHITE      = "FFFFFF"
LIGHT_GREY = "F7F9FC"
MID_GREY   = "E2E8F0"
DARK_GREY  = "4A5568"

# Status fills
G_FILL = PatternFill("solid", fgColor="D4EDDA"); G_FONT = Font(name="Calibri", size=10, color="155724", bold=True)
Y_FILL = PatternFill("solid", fgColor="FFF3CD"); Y_FONT = Font(name="Calibri", size=10, color="856404", bold=True)
R_FILL = PatternFill("solid", fgColor="F8D7DA"); R_FONT = Font(name="Calibri", size=10, color="721C24", bold=True)
O_FILL = PatternFill("solid", fgColor="FFE8D6"); O_FONT = Font(name="Calibri", size=10, color="D35400", bold=True)
B_FILL = PatternFill("solid", fgColor="D1ECF1"); B_FONT = Font(name="Calibri", size=10, color="0C5460", bold=True)

NAVY_FILL  = PatternFill("solid", fgColor=NAVY)
TEAL_FILL  = PatternFill("solid", fgColor=TEAL)
GOLD_FILL  = PatternFill("solid", fgColor=GOLD)
GREY_FILL  = PatternFill("solid", fgColor=LIGHT_GREY)
MID_FILL   = PatternFill("solid", fgColor=MID_GREY)

HDR_FONT   = Font(name="Calibri", size=11, bold=True, color=WHITE)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=WHITE)
REG_FONT   = Font(name="Calibri", size=10)
BOLD_FONT  = Font(name="Calibri", size=10, bold=True)
DARK_FONT  = Font(name="Calibri", size=10, color=DARK_GREY)

THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

def hdr(ws, row, col, val, fill=None, font=None, align="center"):
    c = ws.cell(row=row, column=col, value=val)
    if fill: c.fill = fill
    if font: c.font = font
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = THIN_BORDER
    return c

def cell(ws, row, col, val, fill=None, font=None, halign="left"):
    c = ws.cell(row=row, column=col, value=val)
    if fill: c.fill = fill
    if font: c.font = font
    else: c.font = REG_FONT
    c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=True)
    c.border = THIN_BORDER
    return c

def auto_width(ws, max_w=65):
    for col in ws.columns:
        ml = 0
        cl = get_column_letter(col[0].column)
        for c in col:
            if c.value:
                for ln in str(c.value).split('\n'):
                    ml = max(ml, len(ln))
        ws.column_dimensions[cl].width = min(max(ml + 3, 12), max_w)

# ══════════════════════════════════════════════════════════════════════════════
# 1.  CLIENT DASHBOARD SPEC v2
# ══════════════════════════════════════════════════════════════════════════════
SPEC_HEADERS = [
    "Dashboard Element Type", "Element Name / Attributes",
    "What it Shows (Plain English)", "Target Portal Role / User",
    "Suggested Visualization", "Business Purpose & Decision Impact",
    "Available Filters", "Source KPI Map", "Build Status"
]

# Each sheet: (tab_name, tab_colour, [ (type, name, desc, role, viz, purpose, filters, kpi, status) ])
SHEETS = [

# ── 1. OVERVIEW HUB ───────────────────────────────────────────────────────────
("1. Overview Hub", "1B365D", [
    ("Headline Card","Total Users","Lifetime count of all registered app users with % change from last month.","Founder / Growth","KPI Card","Baseline growth metric for platform scale.","City, Zone, Date","Q19","Ready to Build"),
    ("Headline Card","Active Users","Unique customers who placed ≥1 order in last 30 days.","Founder / Support","KPI Card","Core engagement health signal.","City, Zone, Date","Q19","Ready to Build"),
    ("Headline Card","Inactive Users","Users dormant for 31–60 days with no order activity.","Support Team","KPI Card","Reactivation campaign target pool.","City, Zone","Q26","Ready to Build"),
    ("Headline Card","Churned Users","Users with no order in last 60+ days — at risk of permanent loss.","Support Team","KPI Card","Tracks revenue leakage from lapsed users.","City, Zone","Q26","Ready to Build"),
    ("Headline Card","Total Vendors","Count of active third-party ironing workshop partners.","Founder / Ops","KPI Card","Monitors partner network capacity.","City, Zone","Q40","Ready to Build"),
    ("Headline Card","Total Ironers","Total pressmen currently assigned and active across all stores.","Operations Head","KPI Card","Workforce coverage vs. demand.","City, Zone, Vendor","Q9b","Ready to Build"),
    ("Headline Card","Today's Gross Sales","Total rupee value billed today. Includes all payment modes.","Finance Team","KPI Card","Daily revenue pulse check.","Store, Zone","Q28","Ready to Build"),
    ("Headline Card","Today's Active Orders","Live count of orders across all in-progress states.","Store Manager","KPI Card","Real-time backlog tracker.","Store, Zone","Q9","Ready to Build"),
    ("Headline Card","Today's Completed Sales vs Pending Sales","Two-number card: invoiced sales of completed orders vs. revenue locked in pending orders.","Finance Team","Double KPI Card","Clarifies cash inflow vs. pipeline. 'Lesser orders = more completed sales'.","Store, Zone","Q28","Ready to Build"),
    ("KPI Card","Pending Orders Sales Contribution","Rupee value of all orders currently in Pickup / Ironing / Delivery queue.","Finance Team","KPI Card","Revenue pipeline visibility — pending gross sales.","Store, Zone","Q9, Q28","Ready to Build"),
    ("Table","Store Performance Grid","Ranks all stores by SLA compliance, daily sales, and orders. Columns: Store | Orders | Sales | Pickup SLA% | Delivery SLA% | Complaints | Rating.","Founder","Ranked Table","Supports expansion decisions and low-performer flags.","Zone, Date Range","Q1, Q3, Q40","Ready to Build"),
    ("Table","Delayed Orders Alert (>24h)","Live list of bags stuck in ironing or delivery for more than 24 hours. Columns: Order ID | Store | Stage | Hours Pending | Customer.","Store Manager","Alert Table","Prevents SLA breaches. Escalates critical delays.","Store","N/A","Requires New API: /api/delayed-orders"),
    ("Table","Ironer Performance Summary Table","Breakdown of ironers by level (Level 1, Level 2, Trainee), present count, absent, and attendance % per store. Leaderboard of top ironers.","Operations Head","Table with Rank","Monitors live ironer levels, daily productivity, and earns bonus visibility.","Store, Zone, Date","Q9b","Ready to Build"),
    ("Table","Staff & Rider Attendance Shift Tracker","Logs check-in / check-out times, shift duration, late arrivals, and total working hours per staff and rider.","Ops Head / Logistics","Timeline Grid Table","Shift compliance and manpower cost visibility.","Store, Date, Role","Q18a","Ready to Build"),
    ("Line Chart","Sales Volume vs. Order Size Analysis","Dual-axis chart comparing gross sales (₹) against new order count and average order value over time.","Founder / Finance","Dual-Axis Trend Chart","Pricing strategy guide. Higher orders + lower sales = basket-size problem.","City, Zone, Date Range","Q28, Q30","Ready to Build"),
    ("Line Chart","[SUGGESTION] App Version Order Share","Monitors order share and checkout failures split by app version and OS (Android / iOS). Tracks forced-upgrade impact.","Founder / Tech","Trend Graph","Identifies app bugs causing checkout drop-off per version.","OS Type, App Version, Date","N/A","Ready to Build"),
]),

# ── 2. USER ANALYTICS ─────────────────────────────────────────────────────────
("2. User Analytics", "0D7377", [
    ("Headline Card","Total Users","Total registered users since app launch.","Growth Team","KPI Card","Baseline platform size metric.","City, Date","Q19","Ready to Build"),
    ("Headline Card","Onboarded Users","Users who completed their first order after registration.","Marketing","KPI Card","Tracks onboarding funnel success rate.","City, Zone, Date","Q20","Ready to Build"),
    ("Headline Card","No Address Users","Registered users who have not yet added a delivery address.","Marketing","KPI Card","Identifies drop-off at the address step of onboarding.","City, Zone","Q20","Ready to Build"),
    ("Headline Card","Engaged Users","Users who have placed 2–4 orders (repeat but not habitual).","Growth Team","KPI Card","Middle-of-funnel engagement signal.","City, Zone","Q19","Ready to Build"),
    ("Headline Card","Active Users","Users with ≥1 order in last 30 days.","Growth Team","KPI Card","Core DAU / MAU health metric.","City, Zone, Date","Q19","Ready to Build"),
    ("Headline Card","Value Users","Users with 5–9 lifetime orders — high-value habitual segment.","Growth Team","KPI Card","Priority retention group.","City, Zone","Q19","Ready to Build"),
    ("Headline Card","Churned Users","Users with 0 orders in last 60+ days.","Support Team","KPI Card","Reactivation targeting size.","City, Zone","Q26","Ready to Build"),
    ("Headline Card","Inactive Users","Users dormant 31–60 days — pre-churn warning segment.","Support Team","KPI Card","Early warning before full churn.","City, Zone","Q26","Ready to Build"),
    ("Form Grid","[NEW] Loyalty Program Settings","Configure loyalty reward type, max reward coin value, min redeem coins, expiry days, streak coins, tiers (Basic, Pro, Elite), and recharges rules. Columns: S.NO | Max Reward (2.00%) | Reward Type (Percentage) | Coin Value (1.00) | Min Redeem (25) | Expiry Days (90) | Maximun Redeemtion Percentage (20%) | Completed Order Coins (5) | Streak Coins (10) | Referral (25) | First Order Bonus (10) | App Review (10) | Basic Tier (0 - 50) | Basic Tier Coins Per Order (5) | Basic Tier Badge | Pro Tier (50 - 200) | Pro Tier Coins Per Order (6) | Pro Badge | Elite Tier (200 - 500) | Elite Tier Coins Per Order (7) | Elite Badge | Top-Up Min (1000) | Top-Up Bonus (10) | Earning Rules (T&C) | Status (Active) | Action.","Admin / Marketing","Configuration Form Table","Maintains parameters for the loyalty program coin system.","Tier Name, Status","N/A","Ready to Build"),
    ("Table","[NEW] Loyalty Program Reports","Lifetime Coins (88,152), User Balance (52,643), Spent Coins (29,398), Lost Coins (7,239), Welcome Bonus (7,040), Referrals (10,250), Completed Orders (55,812), Streak (5,600), Top-Up (100), App Reviews (9,350). History Columns: S.No | Transaction Details | Booking ID | User Name | Mobile | Coins | LTC | Active Coins | Reward Source | Status | Transaction Date | Expired Date | Referred By | Referred Date | Zone | Email.","Growth / Finance","Ledger Table with Search","Auditing reward coins and referral transactions.","Zone, Transaction Type, Date Range","Q36a","Ready to Build"),
    ("Line Chart","New User Registration Trend","Daily / Weekly / Monthly count of new user sign-ups plotted as a trend line.","Marketing","Trend Line","Tracks acquisition rate vs. campaign spend.","Zone, Date Range (Day/Week/Month/Year)","Q20","Ready to Build"),
    ("Area Chart","User Retention & Status Trend","Stacked area showing active vs. inactive vs. churned counts over time.","Founder / Growth","Stacked Area Chart","Tells the full user lifecycle health story month by month.","Zone, Date Range","Q19, Q26","Ready to Build"),
    ("Bar Chart","Churned User Personas","Breakdown of churned users split by segment: Onboarded, Engaged, Value, Super Value.","Support Team","Grouped Bar Chart","Identifies which customer tier is churning most — informs win-back offers.","Zone, Date Range","Q26","Ready to Build"),
    ("Bar Chart","Active User Personas","Breakdown of active users split by lifecycle segment.","Growth Team","Grouped Bar Chart","Shows which segment drives daily active orders.","Zone, Date Range","Q19","Ready to Build"),
    ("Line Chart","New vs Old User Orders","Compares order volumes from first-time customers (new) against returning customers each week.","Founder / Marketing","Dual-Line Chart","Measures retention-driven growth vs. pure acquisition.","Zone, Date Range (Day/Week/Month/Year)","Q20, Q26","Ready to Build"),
    ("Line Chart","User vs Revenue vs Orders","Three-line chart showing user growth, revenue trend, and order count on a shared time axis.","Founder","Multi-Line Chart","Shows how user growth translates to revenue — identifies conversion lag.","Zone, Date Range","Q19, Q28, Q43","Ready to Build"),
    ("Line Chart","User vs Orders","Tracks total active user count against total completed orders per period to measure engagement depth.","Growth Team","Dual-Line Chart","Detects falling order frequency among registered users.","Zone, Date Range (Day/Week/Month/Year)","Q19, Q9","Ready to Build"),
    ("Line Chart","Retention & Repeat Rate","Monthly repeat-order rate (% of users placing 2nd+ order within 30 days of first order).","Marketing","Trend Line","Core habit-formation success metric.","Zone, Date Range","Q26, Q27","Ready to Build"),
    ("Grid Table","User Retention Cohort","Month-on-month cohort grid showing what % of users from each acquisition month are still active.","Founder / Growth","Cohort Heat Grid","Identifies which cohort months retained best — validates campaigns.","Zone, Cohort Month","Q26","Ready to Build"),
    ("Bar Chart","Address Adoption Trend","Tracks how many new users added their home address after registration over time.","Marketing","Bar Chart","Identifies friction points in the onboarding address step.","Zone, Date Range","Q20","Ready to Build"),
    ("Bar Chart","Address Conversion Trend","Tracks % of no-address users who eventually add an address and place first order.","Marketing","Conversion Funnel","Measures address nudge campaign effectiveness.","Zone, Date Range","Q20, Q24","Ready to Build"),
    ("Bar Chart","Average Booking Frequency","Average number of orders per active user per month, by store.","Founder","Bar Chart","Tracks habit depth — rising frequency = higher LTV.","Store, Zone, Date","Q19, Q30","Ready to Build"),
    ("Bar Chart","Revenue vs. Orders by Persona","Side-by-side bars comparing revenue contribution and order count for each user persona tier.","Finance Team","Grouped Bar","Sizes each persona's financial weight in the business.","Zone, Date Range","Q28, Q19","Ready to Build"),
    ("Bar Chart","Complaints & Refund by Persona","Complaint volume and refund cost split by user persona — shows which segment raises most issues.","Support Team","Grouped Bar","Identifies high-complaint user tiers to prioritize quality for.","Zone, Date Range","Q42, Q26","Ready to Build"),
    ("Bar Chart","Customer Churn Drop-Off Inactivity Bands","Stacked bars of churned users split into bands: 30–45d, 46–60d, 61–90d, 91–200d, 200+d inactive.","Support / Growth","Stacked Bar Chart","Sizes the re-engagement opportunity at each inactivity milestone.","City, Zone","Q26","Ready to Build"),
    ("Table","User Journey & Stage Migration Timelines","Shows average days customers take to move between lifecycle stages: Onboarded → Active → Value → Super Value.","Growth Team","Timeline Table","Identifies slow conversion stages to accelerate with offers.","Zone, Date Range","N/A","Ready to Build"),
    ("Table","Stage Conversion Funnel (Store-wise)","Per-store breakdown of customer count in each lifecycle stage: Onboarded / Active / Inactive / Churned.","Store Manager","Funnel Table","Standardizes store-level customer health tracking.","City, Zone, Date","Q19, Q26","Ready to Build"),
    ("Table","Persona Complaint & Ticket Breakdown","Table correlating complaint volume and refund amounts for each loyalty persona tier.","Support Team","Table","Flags which customer tier needs most quality investment.","Zone, Date Range","Q42","Ready to Build"),
    ("Bar Chart","[SUGGESTION] Household Size vs. Order Value (AOV)","Compares average order value (garment count) for households with single vs. multiple users.","Marketing / Growth","Scatter / Bar Chart","Tests whether household-targeting increases basket size.","Zone, Community","N/A","Ready to Build"),
    ("Cohort Grid","[SUGGESTION] Subscription vs. One-Time Retention","Compares retention cohorts of subscription users against pay-per-order customers.","Growth Team","Cohort Grid","Validates ROI of subscription plan acquisition.","Zone, Plan Type","N/A","Ready to Build"),
]),

# ── 3. STORE OPERATIONS ───────────────────────────────────────────────────────
("3. Store Operations", "2D6A4F", [
    ("Table","Store-wise Lifecycle Stages","Live pipeline table per store. Columns: Store Name | Total Orders | Upcoming | Received | In Progress | Ironing Completed | Cancelled | Completed | RTO.","Store Manager","Ranked Table","Full operational status at a glance per store.","City, Zone, Date Range (Day/Week/Month/Year)","Q9, Q5, Q17","Ready to Build"),
    ("Table","Upcoming Clothes vs Bag Orders","Forecasts tomorrow's scheduled pickups, split between clothes ironing orders and smart bag orders.","Store Manager","Bar Chart + Table","Improves pre-shift planning and rider allocation.","Store, Date","Q9, Q25","Ready to Build"),
    ("Table","Active Work-In-Progress Orders","Live list of all orders currently in Received / Ironing / Ironing Done stage with bag ID and eta.","Workshop Supervisor","Table","Prevents WIP backlog build-up.","Store, Date","Q2, Q9","Ready to Build"),
    ("Table","Store Bag Inventory","Physical bag audit table: bag barcode, order ID, status (Received / Pending Delivery / RTO), store location.","Store Manager","Inventory Table","Prevents bag loss and misrouting.","Store, Date","Q25","Requires New API: /api/order-inventory"),
    ("KPI Card","Avg Order Processing TAT","Average hours from order received to ironing complete per store per week.","Operations Head","KPI Card + Gauge","Drives workshop speed standards.","Store, Zone, Date","Q2, Q4","Ready to Build"),
    ("Heatmap","Slot Peak Order Flow","Heatmap of order volumes by time slot (6am–10am, 10am–2pm, etc.) per store day.","Store Manager","Time-Slot Heatmap","Optimizes rider and ironer shift timing.","Store, Day, Date Range","Q9a","Ready to Build"),
    ("Alert Badge","Inventory Stock-Out Alerts","Red badge list of stores where chemical or packaging inventory is below the 3-day threshold.","Operations Head","Alert Table","Prevents workshop stoppage from stock-out.","Store","N/A","Requires New API: /api/stock-inventory"),
    ("Table","Daily Stock Ledger","Store-wise daily log of: opening stock, closing stock, stock used, receipts received.","Store Manager","Data Table","Tracks consumable cost and usage efficiency.","Store, Date Range","N/A","Requires New API: /api/stock-inventory"),
    ("KPI Card","Stage-wise Pending TAT Queues","Three KPI cards showing pending hours in: Pickup Stage | Ironing Stage | Delivery Stage.","Operations Team","Triple KPI Cards","Identifies which stage is the current bottleneck.","Store, Zone, Date","Q4","Ready to Build"),
    ("KPI Card","Avg Store-wise Missed TAT SLA","% of orders that exceeded their SLA window per store per week. Ranked from worst to best.","Operations Head","Ranked Table + Card","Holds stores accountable for SLA performance.","Store, Zone, Date","Q4","Ready to Build"),
    ("Table","Store-wise Ironer Performance Matrix","Per-ironer table: Clothes Ironed | Quality Score | Defect Rate | Attendance % | Earnings Estimate.","Operations Head","Leaderboard Table","Individual ironer accountability and bonus eligibility.","Store, Vendor, Date","Q9b, Q18a","Ready to Build"),
    ("Table","Store-wise Rider Performance Matrix","Per-rider table: Deliveries | Acceptance Rate | SLA % | COD Collected | Attendance | Complaints.","Logistics Manager","Leaderboard Table","Rider accountability and payout calculation base.","Store, Zone, Date","Q10, Q12, Q13","Ready to Build"),
    ("Table","Asset Repair & Maintenance Timeline Log","Machine-level log: Machine ID | Type | Last Service | Next Due | Downtime Hours | Repair Cost.","Operations Head","Timeline Table","Prevents surprise machine breakdowns and cost overruns.","Store, Machine Type","N/A","Requires New API: /api/machine-maintenance"),
    ("Table","Inventory Usage Forecasting Ledger","Projects chemical/packaging replenishment dates based on last 7-day usage velocity.","Store Manager","Forecast Table","Proactive restocking before stock-out.","Store, Item Type","N/A","Requires New API: /api/stock-inventory"),
    ("Table","Operations Manager Performance Scorecard","Ranks store managers on: Store SLA, Sales Growth %, User Churn Rate, Team Complaints, Audit Score.","Founder / Ops Head","Ranked Scorecard Table","Objective data-driven manager performance review.","Zone, Date Range","Q40","Ready to Build"),
    ("Table","Smart Bag Cycle & Lifespan Tracker","Tracks bag usage count, average usage before replacement, rotation frequency per customer.","Operations Head","Table","Optimizes bag procurement and customer allocation.","Store, Date","Q25","Ready to Build"),
    ("Table","First-Service SLA Defection Funnel","Tracks churn rate of customers whose first order was delayed or had a damage complaint.","Growth / Quality","Funnel Table","Critical insight: first-bad-experience is the #1 churn driver.","Zone, Date Range","Q24, Q26","Ready to Build"),
    ("Table","[NEW] Asset Category Master","Asset product categories configuration. Columns: S.NO | Unique Code | Product Category | Category Code | Description | Status | Action. Examples: STM-IR-010 (ironers room materials).","Admin","Configuration Table","Configure physical asset classifications.","Status","N/A","Requires New API: /api/asset-management"),
    ("Table","[NEW] Asset Product Type Master","Asset product types setup. Columns: S.NO | Unique Code | Product Type | Category Name | Product Code | Description | Status | Action. Examples: STM-IR-MOS-048 (mosquito net under ironer materials).","Admin / Ops","Configuration Table","Specify asset product type specifications.","Category Name, Status","N/A","Requires New API: /api/asset-management"),
    ("Table","[NEW] Asset Status Master","Define status codes. Columns: S.NO | Status Name | Description | Created At | Asset Status | Action. Examples: Repaired, Condemned.","Admin / Ops","Configuration Table","Status states definition for asset logging.","Status","N/A","Requires New API: /api/asset-management"),
    ("Table","[NEW] Asset Entry Management","Onboard and track assets. Columns: S.No | Asset Code | City | Store / Zone | Asset Category | Asset Type | Serial Number | Price | Purchase Date | Bill Number | Bill Image | Warranty Start/End | AMC Start/End | Status | Change Asset Status | Asset Status Details | Action. Actions: Out for Service, Active. Examples: Cable price 1500, exhaust fan.","Store Manager / Ops","Ledger Table","Track warranties, prices, serial numbers, AMC, bills, and status overrides.","Store, Category, Date","N/A","Requires New API: /api/asset-management"),
    ("Table","[NEW] Asset Service Report","Service repair logs. Columns: S.No | Asset Code | Asset Category | Asset Type | Vendor Name | Zone | Current Status | Previous Status | New Status | Changed By | Replaced By | Replacement Asset Code | Replacement Category | Replacement Type | Taken By | Purpose | Remarks | Status Changed At | Service Start/End Date | Created At.","Ops Head / Finance","Ledger Table","Reconcile repair loops, downtime, replacement assets, and technicians.","Store, Category, Date Range","N/A","Requires New API: /api/asset-management"),
    ("Table","[NEW] Inventory Warehouse Master","Configure warehouses. Columns: S.NO | Name | Location | Manager | Actions. Examples: Warehouse1 (Chennai).","Admin / Ops","Configuration Table","Setup supply nodes and manager mappings.","Location","N/A","Requires New API: /api/inventory-management"),
    ("Table","[NEW] Inventory Stock Management","Warehouse and Store stocks levels. Columns: S.NO | Product | Warehouse | Threshold Qty | Qty | Last Updated | Actions. Highlight threshold warnings (e.g. buttersheet Threshold 25, Qty 75).","Store Manager / Ops","Data Table","Warning lists when rolls, buttersheets, bags, etc. fall below safety levels.","Product, Warehouse","N/A","Requires New API: /api/inventory-management"),
    ("Table","[NEW] Store Stock Transfer Request","Request consumables from warehouse. Columns: S.No | Store | Warehouse | Product | Requested Qty | Approved Qty | Received Qty | Status | Action. Examples: buttersheet request 25 complete.","Store Manager / Ops","Data Table","Initiates transfers and monitors delivery completion.","Store, Product, Status","N/A","Requires New API: /api/inventory-management"),
    ("Table","[NEW] Warehouse Transfer Requests","Incoming store requests at warehouse. Columns: S.NO | Warehouse | Store Zone | Product | Requested Qty | Approved Qty | Received Qty | Status | Actions.","Warehouse Mgr / Ops","Data Table","Fulfill requests and approve dispatches.","Warehouse, Product, Status","N/A","Requires New API: /api/inventory-management"),
]),

# ── 4. RIDER LOGISTICS ────────────────────────────────────────────────────────
("4. Rider Logistics", "C05621", [
    ("Headline Card","Rider Attendance Rate","% of scheduled riders who logged into their shift today.","Logistics Manager","KPI Card","Prevents rider shortage before shift starts.","Zone, Date Range","Q18a","Ready to Build"),
    ("Headline Card","Total Deliveries","Total drop-offs completed by riders in the selected period.","Logistics Team","KPI Card","Volume throughput metric.","Zone, Date Range","Q12","Ready to Build"),
    ("Headline Card","SLA Pickup Compliance","% of doorstep pickups completed within the 2-hour SLA window.","Operations Team","Gauge Chart","Customer trust = pickup on time.","Zone, Date Range","Q1","Ready to Build"),
    ("Headline Card","COD Cash Collected","Total cash in rider wallets awaiting bank deposit. Reconciliation liability.","Logistics Manager","KPI Card","Prevents cash float build-up in field.","Zone","Q13","Ready to Build"),
    ("Table","Rider Performance Grid","Leaderboard: Rider Name | Zone | Deliveries | Acceptance Rate | SLA % | Avg Time | Complaints | COD.","Logistics Manager","Leaderboard Table","Individual rider accountability and payout base.","Rider ID, Zone, Date Range","Q10, Q11, Q12, Q18","Ready to Build"),
    ("Bar Chart","Rider Avg Order Time","Bar chart of average minutes per delivery per rider, sorted by fastest to slowest.","Logistics Manager","Bar Chart","Identifies slow riders for coaching.","Zone, Date Range","Q18","Ready to Build"),
    ("Map Overlay","Rider Route and Heatmap","Geographic heatmap overlaying order density against rider position clusters.","Logistics Head","Map Overlay","Optimizes zone-to-rider assignment efficiency.","Zone, Date","N/A","Requires GPS SDK integration"),
    ("Line Chart","[SUGGESTION] Delivery Cost Per Order","Tracks total logistics cost per delivered order (rider pay + partner fee) over time.","Finance / Logistics","Trend Line","Keeps delivery unit economics in check.","Zone, Date Range","Q17","Ready to Build"),
    ("Table","Rider Shift Logs & Duration Details","Login time, logout time, active session minutes, late clock-in count per rider per shift.","Logistics Manager","Timeline Table","Tracks shift discipline and total productive hours.","Rider, Date","Q18a","Ready to Build"),
    ("Table","Door-to-Door Fulfilment Timeline Tracker","Elapsed time breakdown per order: Assign → Arrive → Pickup Wait → Transit → Drop-off.","Logistics Head","Milestone Table","Finds the exact sub-step causing delivery delays.","Zone, Date Range","Q18","Ready to Build"),
    ("Table","Logistics Performance Breakdown Matrix","Zone-wise table: Accepted | Cancelled | SLA Breached | Avg Delivery Time | COD Compliance.","Logistics Head","Matrix Table","Identifies weak zones for reallocation.","Zone, Rider, Date","Q10, Q11, Q14","Ready to Build"),
    ("Table","Delivery Partner Performance Grid (3P Fleets)","Monitors third-party fleet partners (Pidge, Porter, Dunzo etc.): Orders Assigned | Completed | SLA % | Avg Cost per Task | Cancellation Rate.","Logistics Head / Admin","Leaderboard Table","Controls quality and cost of outsourced last-mile delivery.","Partner, Zone, Date Range","Q10, Q12","Ready to Build"),
]),

# ── 5. FINANCE & REVENUE ─────────────────────────────────────────────────────
("5. Finance & Revenue", "1A535C", [
    ("Headline Card","Gross Sales Value","Total amount billed to customers. Includes delivery fees, taxes, and before discount deduction.","Finance Team","KPI Card","Top-line revenue health.","Store, Date Range","Q28","Ready to Build"),
    ("Headline Card","Net Sales per Store","Revenue per store minus promo/coupon discount costs.","Founder","KPI Card","Per-store profitability baseline.","Store, Date Range","Q29","Ready to Build"),
    ("Headline Card","Average Order Value (AOV)","Average rupee spend per completed order.","Finance Team","KPI Card","Revenue forecasting and basket-size driver.","Store, Date Range","Q30","Ready to Build"),
    ("Headline Card","Wallet Outstanding Liability","Total pre-loaded customer cash wallet balance sitting unused across all accounts.","Finance Team","KPI Card","Tracks deferred liability to customers.","Date","Q36a","Ready to Build"),
    ("Table","[NEW] Expense Category Config","Manage main categories of expenditures. Columns: S.NO | Status Name | Description | Status | Created At | Action.","Admin / CFO","Configuration Table","Chart of accounts categories definition.","Status","N/A","Requires New API: /api/expense-management"),
    ("Table","[NEW] Expense Sub Category Config","Details sub-categories. Columns: S.NO | Category Name | Sub Category Name | Description | Status | Created On | Action.","Admin / CFO","Configuration Table","Micro-spend cost codes definition.","Category Name","N/A","Requires New API: /api/expense-management"),
    ("Table","[NEW] Expense Payment Mode Config","Configure payment modes. Columns: S.NO | Payment Mode Name | Description | Status | Created At | Action.","Finance Head","Configuration Table","Structure payment audit routes.","Status","N/A","Requires New API: /api/expense-management"),
    ("Table","[NEW] Expense Entry Log","Daily expense logs. Columns: S.NO | Date | Store | Category | Sub Category | Amount | Payment Mode | Payment Bill | Remarks | Created On | Action.","Store Manager / Finance","Ledger Table","Logs real-time store operational costs.","Store, Category, Date","N/A","Requires New API: /api/expense-management"),
    ("Table","[NEW] Expense Report","Interactive ledger list. Cards: Total Amount | Total Entries | Avg Per Entry. Columns: S.No | Date | Zone / Store | Category | Sub Category | Amount (₹) | Payment Mode | Bill | Remarks | Created On.","CFO / Finance","Table with Cards","Reconciles store actual spending against receipts.","Store, Category, Date Range","N/A","Requires New API: /api/expense-management"),
    ("Table","[NEW] Expense Analytics Dashboard","Income and expense summary tracker. Cards: Total Income (₹3,343 for 15 completed orders), Total Expense (₹1,553 Promo + Loyalty + Voucher), Promo Expense (₹53 for 1 orders), Loyalty Expense (₹0), Voucher Expense (₹1,500), Net Margin (₹1,790). Tables: Order Details (Order ID | Type | Store | Date | Qty | Base | Promo | Loyalty | Voucher | Payable | Status | Action), Promocode (Order ID | Store | Date | Base Amount | Promo Discount | Final Payable | Status), Loyalty (Order ID | Store | Date | Base Amount | Loyalty Used | Final Payable | Status), Voucher (User | Mobile | Event | Usage | Amount).","Founder / CFO","Dashboard Grid","Full revenue, discounts, and deductions tracker per store and period.","Store, Date Range","Q28, Q29, Q32","Requires New API: /api/expense-analytics"),
    ("Table","[NEW] PIDGE COD Order Report","Courier cash-on-delivery tracking. Columns: S.No | OrderId | Retrigger OrderId | Store Name | Order Date | CDR Id | TRF Id | Delivery Date | Order Amount | Paid Amount | Charges | Balance Amount | Status.","Finance Team","Reconciliation Table","Verifies cash collections from Pidge deliveries against orders.","Store, Payment Status, Date Range","Q10, Q13","Ready to Build"),
    ("Table","[NEW] COD Transaction Report","COD Transfer summary. Columns: Transfer Id | Total Orders | Sum of Order Amount | Sum of Cod Charges | Sum of GST | Sum of Total Cod Charges | Sum of Transfer Amount. Grand Total: ₹ 38,26,746.12, Sum of Transfer: ₹ 36,16,486.","CFO / Finance","Summary Table","Auditing large last-mile cash transfer deposits.","Transfer ID","Q13","Ready to Build"),
    ("Table","[NEW] Wallet History Report","Audit details of customer wallets. Cards: Live (Money Wallet: ₹ 111,476.00, Reward Wallet: ₹ 54,556.00, Expired Reward Voucher: ₹ 320.00), Incoming (Recharge: ₹ 54,900.00, Referral: ₹ 3,900.00, Reward Voucher: ₹ 45,170.00), Refunds (Total: ₹ 1,35,942.00, Partial: ₹ 35,464.00, Ticket: ₹ 1,00,478.00). Columns: S.No | Transaction Details | Transaction Date | User Name | Mobile No | Addition Amount | Deduction Amount | Offer Amount | Paid Amount | Total Amount | Transaction Type | Offer Type | Unique ID | Booking ID | Order Type | Zone Name.","CFO / Finance","Ledger Table with Cards","Reconcile wallet cash flows, credits, refunds, and promo coins.","Zone, Transaction Type, Date Range","Q36a","Ready to Build"),
    ("Line Chart","Order vs Revenue Trend","Historical comparison of order count (left axis) vs. total revenue (right axis) per day/week/month/year.","Finance Team","Dual-Axis Trend","Separates volume growth from revenue growth.","Store, Date Range (Day/Week/Month/Year)","Q28, Q43","Ready to Build"),
    ("Line Chart","Online vs COD Order Trend","Trend comparing digital payment orders vs. cash-on-delivery orders over time.","Finance Team","Dual-Line Chart","Tracks shift toward digital payments — reduces cash risk.","Store, Date Range (Day/Week/Month/Year)","Q31","Ready to Build"),
    ("Line Chart","Online vs COD Revenue Trend","Rupee revenue split: online payments vs. COD collections over time.","Finance Head","Dual-Line Chart","Shows digital revenue share growth vs. cash dependency.","Store, Date Range (Day/Week/Month/Year)","Q31, Q28","Ready to Build"),
    ("Line Chart","Orders vs Clothes vs Revenue","Three-line chart: order count | total garments ironed | total revenue — on same time axis.","Founder / Finance","Multi-Line Chart","Detects basket-size changes (revenue up but clothes down = price shift).","Store, Zone, Date Range","Q9, Q8, Q28","Ready to Build"),
    ("Pie Chart","Payment Mode Split","Donut chart of revenue split: Online (UPI/Card/Wallet) vs. COD vs. Pre-paid Wallet.","Finance Team","Donut Chart","Monitors payment channel health.","Store, Date Range","Q31","Ready to Build"),
    ("Table","COD Reconciliation Status","Store-wise table: Cash Collected by Riders | Deposited | Outstanding | Days Outstanding.","Finance Head","Reconciliation Table","Prevents cash leakage in the field.","Store, Zone, Date Range","Q13","Ready to Build"),
    ("Line Chart","Refund Leakage Trend","Monthly trend of total refund rupees issued, broken down by refund reason.","Finance Head","Trend Chart","Controls cash leak from complaint-driven refunds.","Store, Date Range","Q42","Ready to Build"),
    ("Pie Chart","[SUGGESTION] Wallet Segment Liability Breakout","Donut showing wallet liability split by user segment: Active / Inactive / Churned.","Finance Team","Donut Chart","Sizes refund risk from inactive/churned wallets.","Date","Q36a","Ready to Build"),
    ("KPI Card","[SUGGESTION] Net Profit Margin","Estimated net margin after deducting logistics, labour, and promo costs from net revenue.","Founder / Finance","KPI Card","Business viability monitor.","Store, Date Range","Q28, Q29, Q17","Ready to Build"),
    ("Waterfall","[SUGGESTION] Contribution Margin Per Store","Waterfall chart: Gross Sales → Minus Discounts → Minus Logistics → Minus Labour = Contribution.","CFO","Waterfall Chart","Identifies which stores are truly profitable.","Store, Date Range","Q28, Q17","Ready to Build"),
    ("Table","Prepaid Wallet Recharge & Consumption Ledger","Wallet transaction log: User | Recharge Amount | Consumption Date | Order ID | Balance Remaining.","Finance Team","Data Table","Full wallet money trail for audit and disputes.","User, Date Range","Q36a","Ready to Build"),
    ("Line Chart","Money Wallet Utilised vs Unutilised Split","Trend line comparing the total wallet balance used in orders against unused sitting balance each month.","Finance Head","Dual-Line Chart","Tracks conversion of wallet top-ups into actual orders.","Date Range","Q36a","Ready to Build"),
    ("Bar Chart","Expired Reward Wallet Value","Monthly bar chart of total reward/promo coin value that expired without being redeemed.","Finance Head / Marketing","Bar Chart","Informs reward expiry policy and nudge campaign timing.","Date Range","Q36a","Ready to Build"),
    ("Table","Store Expense Management Ledger","Daily store-level expense log: Date | Store | Category (Chemical/Labour/Utility/Maintenance) | Amount | Approved By.","Finance Head","Data Table","Controls operational cost per store.","Store, Date Range, Category","N/A","Requires New API: /api/stock-inventory"),
]),

# ── 6. GROWTH & MARKETING ────────────────────────────────────────────────────
("6. Growth & Marketing", "553D6B", [
    ("Headline Card","Monthly Active Users (MAU)","Distinct users placing ≥1 order in last 30 days.","Growth / Founder","Trend Graph","Primary growth health metric.","Zone, Date Range","Q19","Ready to Build"),
    ("Headline Card","New Customer Conversion","% of new registrations that converted into a first completed order.","Marketing","Funnel KPI","Acquisition funnel efficiency.","Zone, Date Range","Q24","Ready to Build"),
    ("Table","[NEW] Marketing Notifications Campaign","Campaign table. Filters: Store (-- All Store --), Pincode. Columns: ID | Title | Message | Total Users Sent | Image | Action Text | Store | Pincode | Created At | Action. Interactive: 'Send Notification' button.","Marketing Team","Table with Button","Allows creating campaigns, choosing targeted stores/pincodes, and triggering push blasts.","Store, Pincode","N/A","Ready to Build"),
    ("Table","Promo Code ROI Analysis","Promo vs. organic order split, revenue per promo campaign, cost per acquisition via promo.","CMO","Trend Chart","Stops wasteful discount campaigns.","Promo Code, Date Range","Q23, Q32","Ready to Build"),
    ("Bar Chart","Store CAC (Acquisition Cost)","Avg marketing cost to acquire one paying customer per store.","Marketing","Bar + KPI Card","Measures channel efficiency.","Store, Date Range","Q22","Requires New API: /api/marketing-attribution"),
    ("Line Chart","Promo-to-Organic Retention","Retention rate of promo-first customers vs. organic sign-ups after 30, 60, 90 days.","Marketing","Dual Trend Line","Validates long-term LTV of discounted acquisition.","Zone, Date Range","Q27","Ready to Build"),
    ("Headline Card","Subscription Adoption","% of active users enrolled in a subscription plan (weekly/monthly iron plans).","Growth Team","KPI Card","Tracks predictable recurring revenue.","Zone, Date Range","N/A","Ready to Build"),
    ("Headline Card","Community Orders Rate","% of total orders coming from community/society housing pincode clusters.","Growth Team","KPI Card","Validates community pricing strategy effectiveness.","Zone, Date Range","N/A","Ready to Build"),
    ("KPI Card","[SUGGESTION] Referral Virality K-Factor","Avg referrals sent per user × conversion rate of referred users.","Growth Team","KPI Card","Quantifies organic word-of-mouth growth engine.","Zone, Date Range","N/A","Ready to Build"),
    ("Bar Chart","User-Type Customer Lifetime Value (LTV)","LTV bar per user segment: Onboarded / Active / Value / Super Value — shows revenue per tier.","Finance / Growth","Grouped Bar","Sizes each persona's long-term business value.","Zone, Date Range","Q38","Ready to Build"),
]),

# ── 7. COMPLAINT & QUALITY ────────────────────────────────────────────────────
("7. Complaint & Quality", "922B21", [
    ("Headline Card","Total Complaints Logged","Count of all complaints registered by users in the selected period.","Support Team","KPI Card","Support ticket volume tracker.","Store, Date Range","Q42","Ready to Build"),
    ("Headline Card","SLA Resolved Complaints","% of complaints closed within 48 hours of creation.","Support Team","Gauge Chart","Customer support speed SLA.","Store, Date Range","Q42","Ready to Build"),
    ("Headline Card","Store Audit Compliance","Average quality compliance % scored on monthly store audits.","Quality Team","Gauge Chart","Tracks cleanliness and process standards.","Store, Date Range","Q45a","Ready to Build"),
    ("Table","[NEW] Order Reviews Report","Customer review logs. Columns: S.No | User Name | Email | Mobile | Booking ID | Order Type | Zone | Order Ratings | Review Description | Created At.","Support Team","Data Table","Tracks qualitative feedback ratings and written customer descriptions.","Zone, Rating, Date Range","Q41","Ready to Build"),
    ("Pie Chart","Complaint Type Breakdown","Donut of complaints by category: Iron Damage | Missing Clothes | Late Delivery | Pricing | Other.","Support Team","Donut Chart","Root-cause analysis for complaint reduction.","Store, Date Range","Q42","Ready to Build"),
    ("Table","Pressman Damage Report","Per-pressman table of damage incidents: Ironer Name | Damaged Clothes | Complaint Rate | Refund ₹.","Quality Team","Table","Flags high-error ironers for retraining.","Store, Vendor, Date","Q9b","Ready to Build"),
    ("Line Chart","Staff Complaint Trend","Monthly trend of complaints linked to staff errors (ironer damage, rider attitude, etc.).","Operations Head","Trend Line","Tracks service quality drift over time.","Zone, Date Range","Q42, Q9b","Ready to Build"),
    ("Line Chart","Complaint TAT Trend","Average days to resolve a complaint over time — tracks support team speed improvement.","Support Head","Trend Line","Drives support SLA accountability.","Store, Date Range","Q42","Ready to Build"),
    ("Alert Table","[SUGGESTION] Revenue at Risk / Fraud Alert List","Flags orders with unusually high refund requests, repeat complaints from same user, or promo abuse.","Finance / Support","Alert Table","Prevents refund fraud and bad-actor abuse.","Date Range","Q42","Ready to Build"),
    ("Line Chart","Complaint vs Churn vs RTO Correlation","Three-line chart: monthly complaint count | churn count | RTO orders — reveals quality-churn link.","Founder / Quality","Multi-Line Chart","Proves that complaint spikes directly cause customer churn.","Zone, Date Range","Q42, Q26","Ready to Build"),
    ("Table","Customer Service Executive Performance Scorecard","Agent-wise table: Tickets Handled | Avg Resolution Time | Refunds Given | Refund ₹ | CSAT Score.","Support Head","Leaderboard Table","Drives support team accountability and bonus eligibility.","Agent, Date Range","N/A","Ready to Build"),
]),

# ── 8. TAT ANALYTICS ──────────────────────────────────────────────────────────
("8. TAT Analytics", "1A3A4A", [
    ("Headline Card","Upcoming Pending","Count of orders placed/scheduled where pickup has not started, with % change.","Operations Team","KPI Card + Trend","Prevents pickup backlog.","Store, Zone, Date","Q1","Ready to Build"),
    ("Headline Card","Ironing Pending","Count of orders received at workshop and waiting for ironing to start.","Operations Team","KPI Card + Trend","Keeps ironing queue moving.","Store, Zone, Date","Q2","Ready to Build"),
    ("Headline Card","Delivery Pending","Count of ironed and packed orders waiting for rider dispatch.","Operations Team","KPI Card + Trend","Flags ready orders not yet picked up by rider.","Store, Zone, Date","Q3","Ready to Build"),
    ("Headline Card","RTO Pending","Count of returned/re-triggered orders currently in return loop.","Logistics Team","KPI Card + Trend","Controls return-cycle bottlenecks.","Store, Zone, Date","Q17","Ready to Build"),
    ("Line Chart","Upcoming Pending Trend","Daily trend of upcoming orders count — forecasts next-day logistics needs.","Operations Team","Trend Line","Anticipates rider and slot demand.","Store, Zone, Period","Q1","Ready to Build"),
    ("Line Chart","Ironing Pending Trend","Daily trend of orders waiting in ironing queue per store.","Workshop Supervisor","Trend Line","Surfaces consistent ironing bottleneck patterns.","Store, Vendor, Period","Q2","Ready to Build"),
    ("Line Chart","Delivery Pending Trend","Daily trend of delivery-ready orders not yet dispatched.","Logistics Team","Trend Line","Identifies dispatch gap vs. ironing output.","Store, Zone, Period","Q3","Ready to Build"),
    ("Line Chart","RTO Pending Trend","Monthly trend of return-to-origin orders cycling back through the system.","Logistics Head","Trend Line","Tracks return-process efficiency.","Store, Zone, Period","Q17","Ready to Build"),
    ("Bar Chart","Store-wise TAT Performance","Bar chart ranking stores by avg end-to-end TAT (order placed → delivered). Hours shown.","Operations Head","Bar Chart","Identifies stores consistently missing TAT targets.","Zone, Date Range","Q4","Ready to Build"),
    ("Bar Chart","Ironer Efficiency (Avg TAT)","Bar chart ranking ironers by average hours taken per order from receive to iron-complete.","Workshop Supervisor","Bar Chart","Drives ironing speed standards.","Store, Vendor, Date","Q2, Q9b","Ready to Build"),
    ("Table","Ironer TAT & Volume Explorer","Filterable table: Ironer | Store | Clothes Count | Avg Iron Time (hrs) | Defects | Attendance.","Operations Head","Data Table","Investigates individual ironer efficiency and quality.","Store, Vendor, Date Range","Q9b, Q2","Ready to Build"),
]),

# ── 9. IRONER & STAFF PERFORMANCE ─────────────────────────────────────────────
("9. Ironer & Staff Performance", "2C5F2E", [
    ("Headline Card","Active Ironers","Count of ironers who checked in and completed tasks today.","Operations Head","KPI Card","Daily labour attendance tracking.","Store, Vendor, Date","Q18a","Ready to Build"),
    ("Headline Card","Top Performing Ironer","Name and clothes count of the highest-volume ironer in the zone today.","Operations Head","KPI Card","Flags bonus candidates and peer motivation.","Store, Vendor, Date","Q9b","Ready to Build"),
    ("Headline Card","Avg Time to Iron (Hours)","Average hours from iron-start to iron-complete across all pressmen today.","Operations Head","KPI Card","SLA quality standard monitor.","Store, Vendor, Date","Q9b","Ready to Build"),
    ("Headline Card","Defective Clothes Rate (%)","% of garments flagged as damaged or requiring re-iron during quality checkout.","Quality Team","KPI Card","Damage control metric.","Store, Vendor, Date","Q7","Ready to Build"),
    ("Table","Ironer Performance Scorecard","Leaderboard: Ironer Name | Level (1/2/Trainee) | Clothes Count | Quality Score | Defect Rate | Avg Time | Attendance | Estimated Earnings.","Operations Head","Leaderboard Table","Composite ironer scorecard for bonus and retraining decisions.","Store, Vendor, Date","Q9b, Q18a","Ready to Build"),
    ("Table","Ironer Attendance Details","Daily attendance log: Ironer Name | Store | Shift | Check-In | Check-Out | Working Hours | Status (Present/Absent/Late).","Operations Head","Timeline Table","Shift discipline and manpower cost visibility.","Store, Vendor, Date","Q18a","Ready to Build"),
    ("Table","Ironer Summary Report","Weekly/monthly summary: Total Clothes Ironed | Avg Quality | Total Earnings | Complaints | Absent Days.","Vendor / Ops Head","Summary Table","Vendor-level ironer aggregate performance view.","Vendor, Store, Date Range","Q9b","Ready to Build"),
    ("Table","Rider Leaderboard","Rider ranking: Name | Zone | Total Deliveries | SLA % | COD Compliance | Complaints | Avg Time per Trip.","Logistics Manager","Leaderboard Table","Identifies top riders for bonus and underperformers for coaching.","Zone, Date Range","Q10, Q12, Q13","Ready to Build"),
]),

# ── 10. MASTERS CONFIGURATION ─────────────────────────────────────────────────
("10. Masters Configuration", "424242", [
    ("Form Grid","Community Pricing Master","Configures community rates. Inputs: Community Name, Search Location, Price Per Cloth, MOQ, Radius, Lat/Long, Save Pricing. List Columns: # | Community | Radius | Coordinates | Price | MOQ | Action. Examples: DOSHI GARDENS.","Admin / Growth","Form + Table Grid","Maintains community pricing zones.","Society Name, Zone, Price","N/A","Ready to Build"),
    ("Form Grid","Delivery Master","Configure Delivery Partner dropdown (ADLOGGS/others).","Admin / Ops","Dropdown Interface","Set zone-specific delivery partners.","Zone ID","N/A","Ready to Build"),
    ("Form Grid","Settings Master","Global configurations. Inputs: Minimum Order Quantity (15), Max Order Quantity (50), Bag Cost (50.00), Allowed Quantity Per Bag (25), Vendor Iron Timing (1), Order minimum distance (7), Wallet Discount % (20), Delivery Partner (QWQER), Submit.","Admin","System Config Form","Central toggle panel for operational rules.","Setting Key","N/A","Ready to Build"),
    ("Form Grid","Store Settings","Per-store settings. Columns: S.NO | Zone | VIP Days Limit | VIP Order Limit | FCFS | Cashback | Promocode | Voucher | Wallet | Cloth Image Upload | Referral Wallet Discount % | Ref Min Purchase Amount | Money Wallet Discount % | Complaint Working Hours | Delivery Partner | Ironer IN/OUT Radius | Settings Actions.","Admin / Store Manager","Data Grid Table","Customizes each store's operational parameters.","Store ID","N/A","Ready to Build"),
    ("Form Grid","Partner Priority","Priority mapping: Pickup Partner Priority Management (Select Vendor) | Delivery Partner Priority Management (Select Vendor).","Admin / Logistics","Dropdown Managers","Controls 3P fleets prioritizations.","Zone, Vendor","N/A","Ready to Build"),
    ("Form Grid","Price Master","Configure garment base rates. Inputs: Price Master Name, Cost Per Cloth. Columns: S.No | Price Master Name | Cost Per Cloth | Action. Examples: CHENNAI NEW AREA PRICING (10.00), CHENNAI OLD AREA PRICING (12.00).","Admin / Finance","Form + Table Grid","Garment base pricing templates.","Price Master Name","N/A","Ready to Build"),
    ("Form Grid","Image Upload Master","Configure app images. Inputs: Image Upload (1-5), Description (1-5), Title (1-5), Type dropdown. Columns: ID | Type | Image Upload1 | Image Upload2 | Image Upload3 | Image Upload4 | Image Upload5 | Action. Examples: Welcome Screen, Login Screen.","Admin / Quality","Form + Table Grid","Standardizes uploaded photos categories.","Type","N/A","Ready to Build"),
    ("Form Grid","Pincode Master","Add service pincodes. Inputs: Area Name, Pincode, Price Master selection. Columns: ID | Area Name | Pincode | Price Master | price | Action. Examples: Vadapalani (600055), Villivakkam (600038).","Admin / Ops","Form + Table Grid","Enables pincode mapping and Pricing Master link.","Pincode, Price Master","N/A","Ready to Build"),
    ("Form Grid","Zone Mapping","Maps zones. Inputs: Group Name, Pincode selector (Active/Inactive counts), Price Master, Active Status, Secondary Pincode, Payment Options (COD, Online). Columns: ID | Group Name | Price Master | Group Area Name | Group pincode | Secondary Area Name | Secondary pincode | Status | Action. Examples: Steamee Villivakkam, Steamee - T Nagar.","Admin / Logistics","Form + Table Grid","Maps zones to stores, pricing, and pincodes.","Group Name, Status","N/A","Ready to Build"),
    ("Form Grid","Assign Bag","Bag area mapping. Inputs: Group Name, Pincode selector, Price Master, Active Status, Secondary Pincode, Payment Options. Columns: ID | Group Name | Price Master | Group Area Name | Group pincode | Secondary Area Name | Secondary pincode | Status | Action.","Store Manager","Form + Table Grid","Controls physical bag mapping logic.","Group Name","N/A","Ready to Build"),
    ("Form Grid","Feedback Questions","Question builder. Inputs: Question text, Type dropdown (Text, Dropdown), Options, Add Question button. Columns: # | Question | Type | Options | Action. Examples: How About Ironing Quality (Dropdown: Excellent/Good/Average/Bad).","Admin / Support","Form Builder Table","Maintains NPS and customer surveys.","Question, Type","N/A","Ready to Build"),
    ("Form Grid","[NEW] GST Master","GST tax configuration. Inputs: GST Name, GST Percent, Active Status dropdown. Columns: S.No | Name | Percentage | Active Status | Created Date | Actions. Examples: SERVICE TAX (5%).","Admin / Finance","Form + Table Grid","Maintains tax percentages per transaction.","Name, Active Status","N/A","Ready to Build"),
    ("Form Grid","[NEW] Push Notification Master","Configure push notifications. Inputs: Title, Content, Images, Pincode targets selector. Columns: S.No | Title | Content | Image | Pincodes | Sent Date | Action.","Marketing / Admin","Form + Table Grid","Sends marketing and transactional push notification messages.","Title, Pincode","N/A","Ready to Build"),
    ("Form Grid","[NEW] Communication Master","SMS template schedules configuration. Columns: S.No | No of days | Send Sms (Near Expiry SMS config table for 7days, 6days, 5days...).","Admin / Marketing","Configuration Table","SMS template trigger schedules.","No of days","N/A","Ready to Build"),
    ("Form Grid","[NEW] QR Generation","QR numbers builder. Inputs: Ref (REF-10), Prefix (BAG), Needed Count, Download QR button. Columns: S.No | Ref | Used By | QR Number | QR. Examples: REF-9, Used By: BAG-13429.","Admin / Ops","Form + Table Grid","Generates physical barcodes and QR tags for bags.","Ref, QR Number","N/A","Ready to Build"),
    ("Form Grid","[NEW] Store Creation Master","Store onboarding progress. Form steps: Pincode | Zone | Timeslot | Vendor | Ironer | Settings | TAT. Inputs: Add Pincode (Area Name, Pincode), New Zone Mapping, Timeslot, Iron Vendor, Ironers, Store Settings, TAT Config, Verify Details.","Admin / Founder","Multi-step Wizard","Standardizes launching and configuring new store units.","Pincode, Zone","N/A","Ready to Build"),
]),

# ── 11. OPERATIONAL LISTS & TABLES ────────────────────────────────────────────
("11. Operational Lists & Tables", "5D4037", [
    ("Form Grid","[NEW] Manual Order Form","Manual order entry. Inputs: Select order / cloth order / bag order, Mobile Number, Payable calculator (with savings tracker).","Store Manager","Interactive Form","Creates offline/phone orders.","Customer Mobile","N/A","Ready to Build"),
    ("Table","[NEW] Order List","Core order tracker table. Status tabs (All: 398, Confirmed: 71, Rider Accepted: 20, Rider pickedup: 3, Delivered to Vendor: 110, Transferred: 31, Received: 28, Ironing: 0, RTO: 0, Completed: 35, Hold: 0, Scheduled: 64, Rider Accepted: 5, Rider Pickedup: 2, RTO Delivered: 0, Temp Cancelled: 26). Columns: # | Name | Mobile | Order ID | Bag ID | Pickup Date | Order Date | Status | TAT Status | TAT Timer (e.g. 1d 23h 51m 23s) | Pickup ID | Delivery ID | Action.","Store Manager / Support","Table with Status Tabs","Full order execution pipeline manager.","Status, Store/Zone, Date Range, TAT Status","Q9, Q1, Q3","Ready to Build"),
    ("Table","[NEW] Scheduled Orders / Pickup Hold","Pickup and delivery hold logs. Tabs: Pickup Hold Dashboard, Delivery Hold Dashboard, Standard Hold, Insufficient Funds, Alert. Columns: Order ID | Payment Mode | Transaction Id | User | Mobile | Pickup Date | Pickup Time | Delivery Date | Hold Duration | Status | Action (Cancel, Send to Partner).","Logistics / Ops","Table with Tabs","Manages future scheduled orders and payment-hold order releases.","Store, Date Range","Q9a","Ready to Build"),
    ("Table","[NEW] Completed Orders History","Successful bookings log. Columns: S.No | Order Type | Order ID | User Name | Bag Used | Booking Date | Pickup Date | Pickup ID | Delivery ID | Order Status | Reviews.","Admin / Finance","Table with Reviews Link","Audit past completed sales.","Zone, Date Range","Q28","Ready to Build"),
    ("Table","[NEW] Cancel Orders History","Cancelled bookings log. Columns: S.No | Order Type | Zone | Order ID | User Name | Booking Date | Order Date | Pickup ID | Order Status.","Support / Admin","Table","Tracks cancellations.","Zone, Date Range","Q14","Ready to Build"),
    ("Table","[NEW] Onhold Orders Dashboard","Orders waiting for delivery partner assignment. Cards: Total On Hold (16), Normal (<=1 Day, 0), Warning (2-4 Days, 0), Critical (>4 Days, 16). Columns: # | Order ID | User | Reference | Bag Order Status | On Hold Time | Order Date | Delivery Date | Status | Action (Cancel).","Logistics Manager","Table with Alert Cards","Escalates delivery partner assignments gaps.","Store, Date Range","N/A","Ready to Build"),
    ("Table","[NEW] New Onhold Orders List","New on-hold order queue. Columns: # | Order ID | Payment Mode | Transaction Id | User | Mobile | Delivery Date | Hold Duration | Status | Action.","Store Manager / Support","Table","Logs newly flagged hold orders.","Store, Date Range","N/A","Ready to Build"),
    ("Table","[NEW] Telecalling Status Config","Status definitions. Columns: S.NO | Status Name | Description | Created At | TeleCalling Status | Action (e.g., Prospect, Warm Lead, Disqualified, Call Back, Future Follow Up, RNR).","Telecall Team Lead","Config Table","Setup call lead pipelines.","Status","N/A","Requires New API: /api/telecall-leads"),
    ("Table","[NEW] Telecall Action Taken Config","Action taken options. Columns: S.NO | Action Taken Name | Description | Created At | Action Taken Status | Action (e.g., Follow-up Scheduled, App Usage Explained).","Telecall Team Lead","Config Table","Setup action logging options.","Status","N/A","Requires New API: /api/telecall-leads"),
    ("Table","[NEW] Telecall Disqualified Reason","Disqualification reasons. Columns: S.NO | Reason Name | Description | Created At | TeleCalling Status | Action (e.g., Already Onboarded, Wrong Number, Not Interested).","Telecall Team Lead","Config Table","Disqualification analytics setup.","Status","N/A","Requires New API: /api/telecall-leads"),
    ("Table","[NEW] TeleCall History Report","Outreach histories log. Columns: S.No | User Name | Mobile | Email | Register At | Zone | Follow-up Date | Follow-up Time | Tele Call Status | Action Taken | Priority | Service | Reason | Remarks | Changed By | Call Date.","Telecall Lead / Admin","Table","Audit lead logs, callbacks, outcomes, and agent comments.","Zone, Telecall Status, Date Range","N/A","Requires New API: /api/telecall-leads"),
    ("Table","[NEW] Transfer Order (Bulk)","Bulk transfers interface. Columns: Order ID | Customer Name | Bag ID | Quantity | Status | Delivery Date.","Operations Head","Table","Transfer orders in bulk.","Store, Date Range","N/A","Requires New API: /api/order-transfers"),
    ("Table","[NEW] Batch Transfers List","Batch order transfers. Columns: # | Batch ID | Origin Zone | Processing Zone | Order Count | Bag Count | Created At | Action (Print).","Logistics Head","Table with Action","Tracks batch order dispatches between centers.","Zone, Date Range","N/A","Requires New API: /api/order-transfers"),
    ("Table Grid","Wardrobe (Garment Proof Logs)","Order-wise garment photo log. Columns: Order ID | Vendor | Ironer | Clothes Count | Photos Count | View Gallery.","Quality Team","Table with Image Links","Verifies clothes condition before ironing to settle damage claims.","Vendor, Date Range","N/A","Ready to Build"),
    ("Visual Grid","Rack Management Layout","Visual slot grid (A1–D4) showing occupied/empty rack status, Order ID, and bag barcode per shelf.","Store Manager","Visual Slot Grid Map","Fast workshop storage search and bag check-out speed.","Store, Rack Type","N/A","Ready to Build"),
    ("Chat Box","Customer Support Chat Room","Live in-app chat logs between support agents and customers. Filterable by status: Open/Resolved.","Support Team","Chat Interface","Centralises all customer communication for quick issue resolution.","User, Date, Status","N/A","Ready to Build"),
    ("Table Grid","Zonewise Report List","Zone-level order list. Columns: S.No | Customer Name | Mobile | Order ID | Zone | Cloth Count | Date.","Logistics / Ops","Data Table","Quick reference for zone-level order auditing and rider assignment.","Zone, Date Range","N/A","Ready to Build"),
    ("Table Grid","Offer Report List","Promo-usage report. Columns: Used At | Promocode | User Name | Order ID | Order Amount | Promo Discount | Payable | Zone.","Finance / Marketing","Data Table","Tracks promo redemption for ROI and abuse detection.","Promo Code, Zone, Date","N/A","Ready to Build"),
    ("Table Grid","Vendor Summary Report","Workshop vendor summary. Columns: Vendor | Total Orders | Units Ironed | Total Ironers | Queue | Completed.","Ops Head / Vendor","Summary Table","Holds ironing partners accountable for throughput and queue management.","Vendor, Store, Date","N/A","Ready to Build"),
]),

]  # end SHEETS

# Build v2 workbook
wb2 = openpyxl.Workbook()
wb2.remove(wb2.active)  # remove default sheet

for tab_name, tab_colour, rows in SHEETS:
    ws = wb2.create_sheet(title=tab_name)
    ws.sheet_properties.tabColor = tab_colour

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(SPEC_HEADERS))
    t = ws.cell(row=1, column=1, value=f"STEAMEE — {tab_name.split('. ',1)[1] if '. ' in tab_name else tab_name}")
    t.fill = NAVY_FILL
    t.font = TITLE_FONT
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row
    for ci, h in enumerate(SPEC_HEADERS, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = TEAL_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
    ws.row_dimensions[2].height = 28

    # Data rows
    for ri, r in enumerate(rows, 3):
        is_new_update = "[NEW]" in str(r[1])
        for ci, v in enumerate(r, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = REG_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.fill = GREY_FILL if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

            # Soft highlighted background for new updates
            if is_new_update:
                c.fill = PatternFill("solid", fgColor="FFFDF0") # Very light soft gold tint for row

            # Colour Build Status (col 9)
            if ci == 9:
                v_str = str(v)
                if "Ready" in v_str:
                    c.fill = G_FILL; c.font = G_FONT
                elif "Requires New API" in v_str:
                    c.fill = R_FILL; c.font = R_FONT
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Colour Element Type (col 1)
            if ci == 1:
                v_str = str(v)
                if "SUGGESTION" in str(r[1]):
                    c.fill = PatternFill("solid", fgColor="EBF5FB")
                    c.font = Font(name="Calibri", size=10, color="1A5276", italic=True)

            # Highlight Element Name cell specifically if it's new
            if ci == 2 and is_new_update:
                c.fill = PatternFill("solid", fgColor="FFEFA6") # Warm soft gold highlight
                c.font = Font(name="Calibri", size=10, bold=True, color="A05A00")

        ws.row_dimensions[ri].height = 48

    # Freeze top 2 rows
    ws.freeze_panes = "A3"
    auto_width(ws, max_w=65)

    # Set specific widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 30

wb2.save(r"d:\Steamee\STEAMEE_Client_Dashboard_Spec_v2.xlsx")
wb2.save(r"d:\Steamee\STEAMEE_Client_Dashboard_Spec.xlsx")
print("✅ STEAMEE_Client_Dashboard_Spec_v2.xlsx and STEAMEE_Client_Dashboard_Spec.xlsx rebuilt successfully.")

# ══════════════════════════════════════════════════════════════════════════════
# 2.  GAP ANALYSIS — regenerate
# ══════════════════════════════════════════════════════════════════════════════
import subprocess, sys
result = subprocess.run([sys.executable, r"d:\Steamee\update_gap_analysis.py"], capture_output=True, text=True)
print("Gap Analysis output:", result.stdout.strip() or result.stderr.strip())

# ══════════════════════════════════════════════════════════════════════════════
# 3.  UPDATE KPI v3 SUMMARY COUNTS
# ══════════════════════════════════════════════════════════════════════════════
wb3 = openpyxl.load_workbook(r"d:\Steamee\STEAMEE_KPI_Change_Log_v3.xlsx")
ws_kpi = wb3["KPI MASTER LIST"]
ws_sum = wb3["SUMMARY"]

# Count actual statuses in master list
counts = {"KEEP":0,"NEW":0,"UPDATE":0,"DELETE":0,"PORTAL-USE":0,"PORTAL-ADAPT":0,"PORTAL-GAP":0}
for row in ws_kpi.iter_rows(min_row=2, values_only=True):
    change = str(row[4]).strip().upper() if row[4] else ""
    portal = str(row[13]).strip().upper() if len(row)>13 and row[13] else ""
    if change in counts: counts[change] += 1
    if portal in counts: counts[portal] += 1

# Update summary cells
summary_map = {
    "NEW KPIs Added":    counts["NEW"],
    "KPIs Retained":     counts["KEEP"],
    "KPIs Updated":      counts["UPDATE"],
    "KPIs Deleted":      counts["DELETE"],
    "PORTAL-USE":        counts["PORTAL-USE"],
    "PORTAL-ADAPT":      counts["PORTAL-ADAPT"],
    "PORTAL-GAP":        counts["PORTAL-GAP"],
}
for row in ws_sum.iter_rows():
    for c in row:
        if c.value and isinstance(c.value, str):
            for label, cnt in summary_map.items():
                if label.lower() in str(c.value).lower():
                    # update the count cell in same row
                    adj_cell = ws_sum.cell(row=c.row, column=c.column+1)
                    if isinstance(adj_cell.value, (int, float)):
                        adj_cell.value = cnt

wb3.save(r"d:\Steamee\STEAMEE_KPI_Change_Log_v3.xlsx")
print("✅ STEAMEE_KPI_Change_Log_v3.xlsx SUMMARY counts updated.")
print("\n🎉 All 3 files rebuilt and saved successfully!")
