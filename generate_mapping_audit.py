# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_mapping_audit():
    # File Paths
    f_spec = r"d:\Steamee\STEAMEE_Client_Dashboard_Spec.xlsx"
    f_gap = r"d:\Steamee\STEAMEE_Gap_Analysis.xlsx"
    f_out = r"d:\Steamee\STEAMEE_Reports_Mapping_Audit.xlsx"

    # 1. Parse Dashboard Spec to index KPI -> (Sheet, Element Type, Element Name)
    print("Parsing dashboard spec...")
    wb_spec = openpyxl.load_workbook(f_spec, data_only=True)
    kpi_map = {}
    
    for sheet_name in wb_spec.sheetnames:
        # Ignore new marketing tabs (12 & 13) for mapping the original Reports 1 metrics
        if "Mktg" in sheet_name or "12" in sheet_name or "13" in sheet_name:
            continue
        ws = wb_spec[sheet_name]
        for ri in range(3, ws.max_row + 1):
            row_vals = [cell.value for cell in ws[ri]]
            if len(row_vals) > 7 and row_vals[7]: # Column H is Source KPI Map
                kpis = [k.strip() for k in str(row_vals[7]).split(',')]
                for k in kpis:
                    if k not in kpi_map:
                        kpi_map[k] = []
                    kpi_map[k].append((sheet_name, row_vals[1])) # (Sheet, Element Name)

    # 2. Load Gap Analysis base rows (first 112 rows corresponding to Reports 1)
    print("Reading gap analysis base metrics...")
    wb_gap = openpyxl.load_workbook(f_gap, data_only=True)
    ws_gap = wb_gap['GAP ANALYSIS']
    
    audit_rows = []
    for ri in range(2, 114): # Rows 2 to 113 in Gap Analysis
        row_vals = [cell.value for cell in ws_gap[ri]]
        metric = row_vals[0]
        cat = row_vals[1]
        status = row_vals[2]
        kpi_id = row_vals[3]
        source = row_vals[4]
        note_covered = row_vals[5]
        note_missing = row_vals[6]
        
        # Cross-reference with spec to find where it is placed
        mapped_elements = []
        mapped_sheets = []
        
        if status in ['Covered', 'Partial'] and kpi_id and kpi_id != 'N/A':
            # Split by comma if there are multiple KPIs mapped (e.g. Q9, Q28)
            kpi_list = [k.strip() for k in str(kpi_id).split(',')]
            for k in kpi_list:
                locs = kpi_map.get(k, [])
                for sh, elem in locs:
                    if elem not in mapped_elements:
                        mapped_elements.append(elem)
                    if sh not in mapped_sheets:
                        mapped_sheets.append(sh)
        
        elem_str = ", ".join(mapped_elements) if mapped_elements else "N/A"
        sheet_str = ", ".join(mapped_sheets) if mapped_sheets else "N/A"
        
        audit_rows.append([
            metric, cat, status, kpi_id if kpi_id else "N/A", 
            elem_str, sheet_str, source, note_covered, note_missing
        ])

    # 3. Create the new Audit Workbook
    print("Generating STEAMEE_Reports_Mapping_Audit.xlsx...")
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Reports 1 Mapping Audit"
    ws_out.views.sheetView[0].showGridLines = True
    
    # Stylings
    NAVY = "1B365D"
    TEAL = "0D7377"
    WHITE = "FFFFFF"
    
    NAVY_FILL = PatternFill("solid", fgColor=NAVY)
    HDR_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
    REG_FONT = Font(name="Calibri", size=10)
    BOLD_FONT = Font(name="Calibri", size=10, bold=True)
    
    G_FILL = PatternFill("solid", fgColor="D4EDDA"); G_FONT = Font(name="Calibri", size=10, color="155724", bold=True)
    Y_FILL = PatternFill("solid", fgColor="FFF3CD"); Y_FONT = Font(name="Calibri", size=10, color="856404", bold=True)
    R_FILL = PatternFill("solid", fgColor="F8D7DA"); R_FONT = Font(name="Calibri", size=10, color="721C24", bold=True)
    
    THIN_BORDER = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    # Title Row
    ws_out.merge_cells("A1:I1")
    title_cell = ws_out.cell(row=1, column=1, value="STEAMEE — Client expectations (Reports 1) mapping audit to Dashboard Specs")
    title_cell.fill = NAVY_FILL
    title_cell.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_out.row_dimensions[1].height = 32

    # Headers
    headers = [
        "Client Metric (Reports 1)",
        "Category",
        "Status",
        "KPI ID Map",
        "Mapped Element in Spec",
        "Mapped Tab in Spec",
        "Technical Source / Solution",
        "What is Covered / Notes",
        "Gaps & Action Items (Dev Work)"
    ]
    for ci, h in enumerate(headers, 1):
        c = ws_out.cell(row=2, column=ci, value=h)
        c.fill = PatternFill("solid", fgColor=TEAL)
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
    ws_out.row_dimensions[2].height = 28

    # Populate Data
    for ri, row in enumerate(audit_rows, 3):
        status_val = row[2]
        for ci, val in enumerate(row, 1):
            c = ws_out.cell(row=ri, column=ci, value=val)
            c.font = REG_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.fill = PatternFill("solid", fgColor="F7F9FC") if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            
            # Format Status Column (col 3)
            if ci == 3:
                if val == "Covered":
                    c.fill = G_FILL; c.font = G_FONT
                elif val == "Partial":
                    c.fill = Y_FILL; c.font = Y_FONT
                else:
                    c.fill = R_FILL; c.font = R_FONT
                c.alignment = Alignment(horizontal="center", vertical="center")
                
            # Format KPI ID Column (col 4)
            if ci == 4:
                c.font = BOLD_FONT
                c.alignment = Alignment(horizontal="center", vertical="center")
                
        ws_out.row_dimensions[ri].height = 42

    # Autofit columns
    for col in ws_out.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value and not str(cell.value).startswith('='):
                for line in str(cell.value).split('\n'):
                    max_len = max(max_len, len(line))
        ws_out.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 55)

    # Set custom widths for important long-text columns
    ws_out.column_dimensions['A'].width = 38  # Metric
    ws_out.column_dimensions['B'].width = 18  # Category
    ws_out.column_dimensions['E'].width = 25  # Mapped Element
    ws_out.column_dimensions['F'].width = 25  # Mapped Tab
    ws_out.column_dimensions['H'].width = 45  # Covered Note
    ws_out.column_dimensions['I'].width = 45  # Gap Note

    wb_out.save(f_out)
    print("File saved successfully at:", f_out)

if __name__ == "__main__":
    build_mapping_audit()
